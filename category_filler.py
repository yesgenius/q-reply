#!/usr/bin/env python
"""Question category filler script with answer context and response logging support.

This script automatically fills question categories using GigaChat LLM
based on a predefined category dictionary from an Excel file.
Supports using answers to enhance categorization accuracy and logs raw responses.

The script processes questions from an Excel file, categorizes them using
the get_category_prompt module with optional answer context, and saves
results with detailed logging including raw LLM responses.

Usage:
    python category_filler.py

Configuration:
    Adjust settings at the beginning of the script for IDE execution.
"""

import json
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================================
# CONFIGURATION SECTION - Adjust these settings for IDE execution
# ============================================================================

# File paths
INPUT_FILE = Path("in/QA.xlsx")
OUTPUT_DIR = Path("out")

# Sheet names
SHEET_QA = "QA"
SHEET_CATEGORY = "CATEGORY"
SHEET_LOG_PREFIX = "LOG_CATEGORY"
SHEET_LOG_PARAMS = "LOG_CATEGORY_PARAMS"

# Column indices (1-based for openpyxl)
COL_CATEGORY_NAME = 1  # Column A in CATEGORY sheet
COL_CATEGORY_DESC = 2  # Column B in CATEGORY sheet
COL_QA_CATEGORY = 1  # Column A in QA sheet
COL_QA_QUESTION = 2  # Column B in QA sheet
COL_QA_ANSWER = 3  # Column C in QA sheet

# Categorization configuration
USE_ANSWER_FOR_CATEGORIZATION = (
    True  # Include answer in categorization for better accuracy
)
MAX_RETRY_ATTEMPTS = 5  # Maximum number of retry attempts on error
RETRY_DELAY = 2  # Delay between retries in seconds (if needed)

# Logging configuration
LOG_CATEGORY = True  # Set to False to disable category logging sheet
LOG_TO_FILE = True  # Enable logging to text file
LOG_LEVEL = logging.INFO
# LOG_LEVEL = logging.DEBUG

# Processing configuration
START_ROW = 2  # Skip header row
SAVE_FREQUENCY = 5  # Save file every N processed rows (reduced for safety)

# Resume configuration
RESUME_FILE = Path(".category_filler_resume.json")  # Hidden file for resume state

# ============================================================================
# END OF CONFIGURATION SECTION
# ============================================================================


def setup_logging(log_file: Path | None = None) -> None:
    """Configure logging for all modules used by this script.

    Sets up unified logging configuration for the main script and all imported
    modules to ensure consistent formatting and output handling.

    Args:
        log_file: Optional path to log file. If provided, logs will be written
            to both console and file.
    """
    # Create formatter with consistent format
    formatter = logging.Formatter(
        "[%(asctime)s][%(name)s][%(levelname)s][%(filename)s:%(lineno)d][%(message)s]"
    )

    # Get root logger to configure all loggers
    root_logger = logging.getLogger()
    root_logger.setLevel(LOG_LEVEL)

    # Remove any existing handlers to avoid duplicates
    root_logger.handlers.clear()

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(LOG_LEVEL)
    console_handler.setFormatter(formatter)
    root_logger.addHandler(console_handler)

    # File handler if requested
    if log_file and LOG_TO_FILE:
        try:
            log_file.parent.mkdir(parents=True, exist_ok=True)
            file_handler = logging.FileHandler(log_file, encoding="utf-8")
            file_handler.setLevel(LOG_LEVEL)
            file_handler.setFormatter(formatter)
            root_logger.addHandler(file_handler)

            # Log to confirm file logging is active
            logger = logging.getLogger(__name__)
            logger.info(f"Logging to file: {log_file}")
        except Exception as e:
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not create log file: {e}")


# Import the categorization module AFTER setting up logging
# This ensures the module inherits our logging configuration
try:
    from prompts import get_category_prompt
except ImportError:
    print("Error: Could not import get_category_prompt module")
    print("Ensure the module is in the correct path: prompts/get_category_prompt.py")
    sys.exit(1)


def format_json_for_excel(data: Any) -> str:
    """Format JSON data for Excel cell display.

    Formats JSON data (messages or raw responses) for readable display
    in Excel cells while preserving the complete data structure.

    Args:
        data: JSON-serializable data (dict, list, or already stringified JSON).

    Returns:
        Formatted JSON string suitable for Excel cell display.
    """
    try:
        # Handle None or empty data
        if data is None:
            return ""

        # If already a string, try to parse and re-format
        if isinstance(data, str):
            # Check if it's already JSON-formatted
            try:
                parsed_data = json.loads(data)
                data = parsed_data
            except (json.JSONDecodeError, TypeError):
                # Not JSON or parsing failed, return cleaned string
                # Remove excessive whitespace while preserving structure
                lines = data.split("\n")
                cleaned_lines = [line.rstrip() for line in lines]
                return "\n".join(cleaned_lines)

        # Format with indentation and ensure_ascii=False for readability
        formatted = json.dumps(data, ensure_ascii=False, indent=2)

        # Replace escaped newlines with actual newlines for Excel
        formatted = formatted.replace("\\n", "\n")

        # Replace escaped quotes for better readability
        formatted = formatted.replace('\\"', '"')

        return formatted

    except (TypeError, ValueError) as e:
        # If formatting fails, return string representation
        logger = logging.getLogger(__name__)
        logger.debug(f"JSON formatting failed: {e}")
        return str(data)


# Get logger for this module
logger = logging.getLogger(__name__)


class CategoryFiller:
    """Main class for filling question categories using LLM with answer context and response logging."""

    def __init__(self):
        """Initialize the category filler."""
        self.workbook: Workbook | None = None
        self.output_file: Path | None = None
        self.categories: dict[str, str] = {}
        self.processed_count: int = 0
        self.resume_state: dict = {}
        self.timestamp: str = ""

    def validate_input_file(self) -> bool:
        """Validate that input file exists and has required sheets.

        Also checks if the file is accessible (not locked by another program).

        Returns:
            True if validation passes, False otherwise.
        """
        if not INPUT_FILE.exists():
            logger.error(f"Input file not found: {INPUT_FILE}")
            return False

        # Check if file is accessible (not locked)
        try:
            with open(INPUT_FILE, "rb") as f:
                pass  # Just open and close - if this works, file is not locked
        except PermissionError:
            logger.error(f"Cannot access file: {INPUT_FILE}")
            logger.error("The file might be open in Excel or another program.")
            logger.error("Please close the file and run the script again.")
            return False
        except OSError as e:
            logger.error(f"Cannot read file {INPUT_FILE}: {e}")
            return False

        try:
            wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)

            # Check required sheets
            if SHEET_CATEGORY not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_CATEGORY}' not found")
                wb.close()
                return False

            if SHEET_QA not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_QA}' not found")
                wb.close()
                return False

            wb.close()
            return True

        except PermissionError:
            logger.error(f"File is locked: {INPUT_FILE}")
            logger.error(
                "Please close the file in Excel or other programs and try again."
            )
            return False
        except Exception as e:
            logger.error(f"Error reading input file: {e}")
            return False

    def load_categories(self, sheet: Worksheet) -> dict[str, str]:
        """Load category dictionary from CATEGORY sheet.

        Args:
            sheet: The CATEGORY worksheet.

        Returns:
            Dictionary mapping category names to descriptions.

        Raises:
            ValueError: If categories have missing data.
        """
        categories = {}
        empty_rows = 0

        for row_idx in range(START_ROW, sheet.max_row + 1):
            category_name = sheet.cell(row=row_idx, column=COL_CATEGORY_NAME).value
            category_desc = sheet.cell(row=row_idx, column=COL_CATEGORY_DESC).value

            # Skip completely empty rows
            if category_name is None and category_desc is None:
                empty_rows += 1
                continue

            # Validate completeness - both fields must be filled
            if category_name is None or category_desc is None:
                raise ValueError(
                    f"Incomplete category data at row {row_idx} in CATEGORY sheet. "
                    f"Category name: {category_name}, Description: {category_desc}. "
                    f"Both columns A and B must be filled."
                )

            categories[str(category_name).strip()] = str(category_desc).strip()

        if not categories:
            raise ValueError(
                "No categories found in CATEGORY sheet. Please fill columns A and B."
            )

        logger.info(f"Loaded {len(categories)} categories from CATEGORY sheet")
        return categories

    def validate_qa_data(self, sheet: Worksheet) -> list[int]:
        """Validate QA sheet data completeness and identify rows needing processing.

        Args:
            sheet: The QA worksheet.

        Returns:
            List of row indices that need category processing.

        Raises:
            ValueError: If questions or answers are missing.
        """
        rows_to_process = []
        total_rows = 0

        for row_idx in range(START_ROW, sheet.max_row + 1):
            question = sheet.cell(row=row_idx, column=COL_QA_QUESTION).value
            answer = sheet.cell(row=row_idx, column=COL_QA_ANSWER).value
            category = sheet.cell(row=row_idx, column=COL_QA_CATEGORY).value

            # Skip completely empty rows
            if question is None and answer is None and category is None:
                continue

            total_rows += 1

            # Validate that both question and answer are present
            if question is None or answer is None:
                raise ValueError(
                    f"Incomplete QA data at row {row_idx} in QA sheet. "
                    f"Question (column B): {'Present' if question else 'Missing'}, "
                    f"Answer (column C): {'Present' if answer else 'Missing'}. "
                    f"Both columns B and C must be filled."
                )

            # Handle multi-line categories - treat as string and check each line
            if category is not None:
                category_str = str(category).strip()
                # Check if it's a multi-line category
                if "\n" in category_str:
                    # Split and check if ALL lines are valid categories
                    category_lines = [
                        line.strip()
                        for line in category_str.split("\n")
                        if line.strip()
                    ]
                    all_valid = all(cat in self.categories for cat in category_lines)
                    if not all_valid:
                        rows_to_process.append(row_idx)
                        logger.debug(
                            f"Row {row_idx}: Multi-line category contains invalid entries, needs processing"
                        )
                elif category_str == "":
                    rows_to_process.append(row_idx)
                    logger.debug(f"Row {row_idx}: Category is empty, needs processing")
                elif category_str not in self.categories:
                    rows_to_process.append(row_idx)
                    logger.debug(
                        f"Row {row_idx}: Category '{category_str}' not in dictionary, needs processing"
                    )
            else:
                rows_to_process.append(row_idx)
                logger.debug(f"Row {row_idx}: Category is None, needs processing")

        logger.info(f"Validated {total_rows} QA rows")
        return rows_to_process

    def create_log_params_sheet(self, wb: Workbook) -> None:
        """Create or update LOG_CATEGORY_PARAMS sheet with model parameters.

        Args:
            wb: Workbook to add the sheet to.
        """
        if not LOG_CATEGORY:
            return

        # Remove existing sheet if present
        if SHEET_LOG_PARAMS in wb.sheetnames:
            wb.remove(wb[SHEET_LOG_PARAMS])

        # Create new sheet
        params_sheet = wb.create_sheet(SHEET_LOG_PARAMS)

        # Add headers
        params_sheet.cell(row=1, column=1, value="Parameter")
        params_sheet.cell(row=1, column=2, value="Value")
        params_sheet.cell(row=1, column=3, value="Description")

        # Get model parameters
        model_params = getattr(get_category_prompt, "params", {})

        # Configuration parameters
        row = 2
        params_sheet.cell(row=row, column=1, value="--- Configuration Settings ---")
        params_sheet.cell(row=row, column=2, value="")
        params_sheet.cell(row=row, column=3, value="Script configuration parameters")
        row += 1

        config_params = [
            ("TIMESTAMP", self.timestamp, "Processing start time"),
            (
                "USE_ANSWER_FOR_CATEGORIZATION",
                USE_ANSWER_FOR_CATEGORIZATION,
                "Whether answers are used for categorization",
            ),
            (
                "MAX_RETRY_ATTEMPTS",
                MAX_RETRY_ATTEMPTS,
                "Maximum retry attempts on error",
            ),
            ("RETRY_DELAY", RETRY_DELAY, "Delay between retries in seconds"),
            ("LOG_CATEGORY", LOG_CATEGORY, "Enable category logging sheet"),
            ("LOG_TO_FILE", LOG_TO_FILE, "Enable logging to text file"),
            ("LOG_LEVEL", logging.getLevelName(LOG_LEVEL), "Logging verbosity level"),
            ("START_ROW", START_ROW, "First row to process (skip headers)"),
            ("SAVE_FREQUENCY", SAVE_FREQUENCY, "Save file every N rows"),
        ]

        for param_name, param_value, param_desc in config_params:
            params_sheet.cell(row=row, column=1, value=param_name)
            params_sheet.cell(row=row, column=2, value=str(param_value))
            params_sheet.cell(row=row, column=3, value=param_desc)
            row += 1

        # File paths section
        row += 1
        params_sheet.cell(row=row, column=1, value="--- File Paths ---")
        params_sheet.cell(row=row, column=2, value="")
        params_sheet.cell(row=row, column=3, value="Input/output file locations")
        row += 1

        file_params = [
            ("INPUT_FILE", str(INPUT_FILE), "Source Excel file path"),
            ("OUTPUT_DIR", str(OUTPUT_DIR), "Output directory path"),
            ("RESUME_FILE", str(RESUME_FILE), "Resume state file path"),
        ]

        for param_name, param_value, param_desc in file_params:
            params_sheet.cell(row=row, column=1, value=param_name)
            params_sheet.cell(row=row, column=2, value=param_value)
            params_sheet.cell(row=row, column=3, value=param_desc)
            row += 1

        # Model parameters section
        if model_params:
            row += 1
            params_sheet.cell(row=row, column=1, value="--- Model Parameters ---")
            params_sheet.cell(row=row, column=2, value="")
            params_sheet.cell(row=row, column=3, value="LLM configuration settings")
            row += 1

            param_descriptions = {
                "model": "LLM model name",
                "temperature": "Randomness of responses (0-1)",
                "top_p": "Nucleus sampling threshold",
                "stream": "Whether to stream responses",
                "max_tokens": "Maximum tokens in response",
                "repetition_penalty": "Penalty for repeated tokens",
            }

            for param_name, param_value in model_params.items():
                params_sheet.cell(row=row, column=1, value=param_name)
                params_sheet.cell(row=row, column=2, value=str(param_value))
                params_sheet.cell(
                    row=row, column=3, value=param_descriptions.get(param_name, "")
                )
                row += 1

        # Auto-adjust column widths
        for column in params_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            params_sheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Created '{SHEET_LOG_PARAMS}' sheet with model parameters")

    def create_output_file(self) -> tuple[Workbook, Path]:
        """Create output file with timestamp.

        Returns:
            Tuple of (workbook, output_path).
        """
        # Ensure output directory exists
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Generate timestamp
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = OUTPUT_DIR / f"QA_{self.timestamp}.xlsx"

        # Setup logging with file output
        if LOG_TO_FILE:
            log_file = OUTPUT_DIR / f"QA_{self.timestamp}.log"
            setup_logging(log_file)
        else:
            setup_logging()

        logger.info(f"Copying {INPUT_FILE} to {output_file}")

        # Copy file preserving everything
        shutil.copy2(INPUT_FILE, output_file)
        logger.info("File copied successfully")

        # Open the copied file for modifications
        wb = openpyxl.load_workbook(output_file)

        # Handle LOG_CATEGORY sheet based on configuration
        log_sheet_name = SHEET_LOG_PREFIX

        if LOG_CATEGORY:
            # Remove existing LOG_CATEGORY sheet if it exists
            if log_sheet_name in wb.sheetnames:
                logger.info(
                    f"Removing existing '{log_sheet_name}' sheet for fresh start"
                )
                wb.remove(wb[log_sheet_name])

            # Create new LOG_CATEGORY sheet
            log_sheet = wb.create_sheet(log_sheet_name)

            # Add headers to LOG_CATEGORY sheet
            headers = [
                "Original Category",
                "Question",
                "Assigned Category",
                "Confidence",
                "Reasoning",
                "Messages",
                "Response",
            ]
            for col_idx, header in enumerate(headers, 1):
                log_sheet.cell(row=1, column=col_idx, value=header)

            logger.info(f"Created '{log_sheet_name}' sheet with headers")

            # Create LOG_CATEGORY_PARAMS sheet
            self.create_log_params_sheet(wb)
        else:
            # Remove sheets if LOG_CATEGORY is False
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing '{log_sheet_name}' sheet (LOG_CATEGORY=False)")
                wb.remove(wb[log_sheet_name])
            if SHEET_LOG_PARAMS in wb.sheetnames:
                logger.info(f"Removing '{SHEET_LOG_PARAMS}' sheet (LOG_CATEGORY=False)")
                wb.remove(wb[SHEET_LOG_PARAMS])

        # Save modifications
        wb.save(output_file)
        logger.info(f"Created output file: {output_file}")

        return wb, output_file

    def save_resume_state(self, output_file: Path, last_row: int) -> None:
        """Save resume state to file.

        Args:
            output_file: Path to the output file being processed.
            last_row: Last successfully processed row number.
        """
        state = {
            "output_file": str(output_file),
            "last_row": last_row,
            "timestamp": datetime.now().isoformat(),
        }

        try:
            with open(RESUME_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2, ensure_ascii=False)
            logger.debug(f"Saved resume state at row {last_row}")
        except Exception as e:
            logger.warning(f"Could not save resume state: {e}")

    def load_resume_state(self) -> dict | None:
        """Load resume state if exists.

        Returns:
            Resume state dictionary or None if not found.
        """
        if not RESUME_FILE.exists():
            return None

        try:
            with open(RESUME_FILE, encoding="utf-8") as f:
                state: dict = json.load(f)

            # Validate output file still exists
            output_path = Path(state["output_file"])
            if output_path.exists():
                logger.info(f"Found resume state from {state['timestamp']}")
                logger.info(f"Will continue from row {state['last_row'] + 1}")
                return state
            logger.warning("Resume output file not found, starting fresh")
            return None

        except Exception as e:
            logger.warning(f"Could not load resume state: {e}")
            return None

    def clear_resume_state(self) -> None:
        """Clear resume state file."""
        if RESUME_FILE.exists():
            try:
                RESUME_FILE.unlink()
                logger.debug("Cleared resume state")
            except Exception as e:
                logger.warning(f"Could not clear resume state: {e}")

    def categorize_question_with_retry(
        self, question: str, answer: str | None = None
    ) -> dict[str, Any]:
        """Categorize a question with retry logic on errors.

        Args:
            question: The question text to categorize.
            answer: Optional answer text to provide additional context.

        Returns:
            Dictionary with category, confidence, reasoning, messages, and response.
        """
        import time

        last_error = None

        for attempt in range(1, MAX_RETRY_ATTEMPTS + 1):
            try:
                # Run categorization - returns three values
                result_json, messages_list, raw_response = get_category_prompt.run(
                    question,
                    answer=answer if USE_ANSWER_FOR_CATEGORIZATION and answer else None,
                )
                result: dict[str, Any] = json.loads(result_json)

                # Format messages and response for Excel readability
                result["messages"] = format_json_for_excel(messages_list)
                result["response"] = format_json_for_excel(raw_response)

                if attempt > 1:
                    logger.info(f"Categorization succeeded on attempt {attempt}")

                return result

            except Exception as e:
                last_error = e
                logger.warning(
                    f"Categorization attempt {attempt}/{MAX_RETRY_ATTEMPTS} failed: {e}"
                )

                if attempt < MAX_RETRY_ATTEMPTS:
                    # Add delay between retries if configured
                    if RETRY_DELAY > 0:
                        time.sleep(RETRY_DELAY)
                    continue

        # All attempts failed
        logger.error(f"All {MAX_RETRY_ATTEMPTS} categorization attempts failed")
        return {
            "category": "Error",
            "confidence": 0.0,
            "reasoning": f"All retry attempts failed. Last error: {last_error!s}",
            "messages": "[]",
            "response": "{}",
        }

    def process_row(
        self, sheet_qa: Worksheet, sheet_log: Worksheet | None, row_idx: int
    ) -> bool:
        """Process a single row from QA sheet.

        Args:
            sheet_qa: The QA worksheet.
            sheet_log: The LOG_CATEGORY worksheet (if enabled).
            row_idx: Row index to process.

        Returns:
            True if processing was successful, False otherwise.
        """
        try:
            # Get data from QA sheet
            current_category = sheet_qa.cell(row=row_idx, column=COL_QA_CATEGORY).value
            question = sheet_qa.cell(row=row_idx, column=COL_QA_QUESTION).value
            answer = sheet_qa.cell(row=row_idx, column=COL_QA_ANSWER).value

            if question is None:
                logger.warning(f"Empty question at row {row_idx}, skipping")
                return True

            # Log processing start
            question_preview = (
                str(question)[:80] + "..." if len(str(question)) > 80 else str(question)
            )
            logger.info(f"Processing row {row_idx}: {question_preview}")

            # Log if using answer context
            if USE_ANSWER_FOR_CATEGORIZATION and answer:
                answer_preview = (
                    str(answer)[:50] + "..." if len(str(answer)) > 50 else str(answer)
                )
                logger.debug(f"Using answer for categorization: {answer_preview}")

            # Categorize the question with retry logic
            result = self.categorize_question_with_retry(
                str(question), str(answer) if answer else None
            )

            # Update category in QA sheet
            sheet_qa.cell(row=row_idx, column=COL_QA_CATEGORY, value=result["category"])

            # Log to LOG_CATEGORY sheet if enabled
            if LOG_CATEGORY and sheet_log:
                # Write to the same row number as in QA sheet for alignment
                sheet_log.cell(row=row_idx, column=1, value=current_category or "")
                sheet_log.cell(row=row_idx, column=2, value=question)
                sheet_log.cell(row=row_idx, column=3, value=result["category"])
                sheet_log.cell(row=row_idx, column=4, value=result["confidence"])
                sheet_log.cell(row=row_idx, column=5, value=result.get("reasoning", ""))
                sheet_log.cell(row=row_idx, column=6, value=result.get("messages", ""))
                sheet_log.cell(row=row_idx, column=7, value=result.get("response", ""))

            # Log result
            logger.info(
                f"Row {row_idx} processed: '{result['category']}' "
                f"(confidence: {result['confidence']:.2f})"
            )

            self.processed_count += 1
            return True

        except Exception as e:
            logger.error(f"Error processing row {row_idx}: {e}")
            return False

    def run(self) -> bool:
        """Main execution method.

        Returns:
            True if execution completed successfully, False otherwise.
        """
        try:
            # Setup initial logging to console only
            setup_logging()

            logger.info("=" * 70)
            logger.info("Starting Category Filler Script")
            logger.info(
                f"Answer Usage: {'ENABLED' if USE_ANSWER_FOR_CATEGORIZATION else 'DISABLED'}"
            )
            logger.info(f"Max Retry Attempts: {MAX_RETRY_ATTEMPTS}")
            logger.info("=" * 70)

            # Check for resume state
            resume_state = self.load_resume_state()

            if resume_state:
                # Resume from previous run
                logger.info("Resuming from previous run...")
                self.output_file = Path(resume_state["output_file"])

                # Extract timestamp from filename for log file
                filename = self.output_file.stem  # QA_YYYY-MM-DD_HHMMSS
                if filename.startswith("QA_"):
                    self.timestamp = filename[3:]  # Remove "QA_" prefix

                # Setup logging with file for resumed session
                if LOG_TO_FILE:
                    log_file = OUTPUT_DIR / f"QA_{self.timestamp}.log"
                    setup_logging(log_file)
                    logger.info("=" * 70)
                    logger.info("RESUMED SESSION")
                    logger.info("=" * 70)

                self.workbook = openpyxl.load_workbook(self.output_file)
                start_from_row = resume_state["last_row"] + 1

                # Load categories from existing file
                if SHEET_CATEGORY not in self.workbook.sheetnames:
                    logger.error(f"Sheet '{SHEET_CATEGORY}' not found in resume file")
                    return False

                sheet_category = self.workbook[SHEET_CATEGORY]
                self.categories = self.load_categories(sheet_category)

            else:
                # Fresh start
                logger.info("Starting fresh processing...")

                # Step 1: Validate input file
                logger.info("Step 1: Validating input file...")
                if not self.validate_input_file():
                    logger.error("=" * 70)
                    logger.error("FAILED: Cannot proceed with processing")
                    logger.error("=" * 70)
                    return False

                # Step 2: Create output file (this also sets up file logging)
                logger.info("Step 2: Creating output file...")
                self.workbook, self.output_file = self.create_output_file()

                # Step 3: Load and validate categories
                logger.info("Step 3: Loading categories from CATEGORY sheet...")
                sheet_category = self.workbook[SHEET_CATEGORY]
                try:
                    self.categories = self.load_categories(sheet_category)
                except ValueError as e:
                    logger.error(f"Category validation failed: {e}")
                    logger.error(
                        "Please ensure all categories in column A have descriptions in column B"
                    )
                    return False

                start_from_row = START_ROW

            # Display loaded categories
            logger.info(f"Categories loaded: {', '.join(self.categories.keys())}")

            # Step 4: Initialize categorization system
            logger.info("Step 4: Initializing GigaChat categorization system...")

            # Initialize with categories
            get_category_prompt.update_system_prompt(categories=self.categories)
            logger.info("Categorization system initialized successfully")

            if USE_ANSWER_FOR_CATEGORIZATION:
                logger.info("Note: Answers will be used for enhanced categorization")

            # Step 5: Validate QA data
            logger.info("Step 5: Validating QA sheet data...")
            sheet_qa = self.workbook[SHEET_QA]

            try:
                rows_to_process = self.validate_qa_data(sheet_qa)
            except ValueError as e:
                logger.error(f"QA data validation failed: {e}")
                logger.error(
                    "Please ensure all questions (column B) and answers (column C) are filled"
                )
                return False

            # Filter rows based on resume state
            if resume_state:
                original_count = len(rows_to_process)
                rows_to_process = [r for r in rows_to_process if r >= start_from_row]
                logger.info(
                    f"Resuming: skipping {original_count - len(rows_to_process)} already processed rows"
                )

            if not rows_to_process:
                logger.info(
                    "No rows need processing. All categories are already filled correctly."
                )
                self.clear_resume_state()
                return True

            logger.info(
                f"Found {len(rows_to_process)} rows that need category assignment"
            )

            # Get LOG_CATEGORY sheet if enabled
            sheet_log = None
            if LOG_CATEGORY:
                if SHEET_LOG_PREFIX in self.workbook.sheetnames:
                    sheet_log = self.workbook[SHEET_LOG_PREFIX]
                else:
                    logger.warning(
                        f"LOG_CATEGORY is enabled but '{SHEET_LOG_PREFIX}' sheet not found"
                    )

            # Step 6: Process rows
            logger.info("Step 6: Processing rows...")
            if USE_ANSWER_FOR_CATEGORIZATION:
                logger.info("Note: Using answers for enhanced categorization accuracy")
            logger.info("-" * 70)

            successfully_processed: list[int] = []  # Track successfully processed rows

            for idx, row_idx in enumerate(rows_to_process, 1):
                # Process row
                success = self.process_row(sheet_qa, sheet_log, row_idx)

                if not success:
                    logger.error(f"Failed to process row {row_idx}")

                    # Emergency save after error
                    if successfully_processed:
                        try:
                            self.workbook.save(self.output_file)
                            last_successful = successfully_processed[-1]
                            self.save_resume_state(self.output_file, last_successful)
                            logger.info(
                                f"Emergency save after error: saved up to row {last_successful}"
                            )
                        except Exception as save_error:
                            logger.error(
                                f"Could not perform emergency save: {save_error}"
                            )

                    continue  # Skip to next row

                # Track successful processing
                successfully_processed.append(row_idx)

                # Periodic save based on SAVE_FREQUENCY
                if idx % SAVE_FREQUENCY == 0:
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(
                            f"Progress saved: {idx}/{len(rows_to_process)} rows processed "
                            f"(saved up to row {row_idx})"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save progress: {save_error}")

                # Save on the last row
                if idx == len(rows_to_process):
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(
                            f"Final batch saved: {idx}/{len(rows_to_process)} rows completed"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save final batch: {save_error}")

            # Final save
            self.workbook.save(self.output_file)
            logger.info(f"Final save completed: {self.output_file}")

            # Clear resume state on successful completion
            self.clear_resume_state()

            # Summary
            logger.info("-" * 70)
            logger.info("Processing completed successfully!")
            logger.info(f"Total rows processed: {self.processed_count}")
            logger.info(
                f"Answer usage was: {'ENABLED' if USE_ANSWER_FOR_CATEGORIZATION else 'DISABLED'}"
            )
            logger.info(f"Output file: {self.output_file}")
            if LOG_CATEGORY and sheet_log:
                logger.info(f"Detailed logs saved in '{SHEET_LOG_PREFIX}' sheet")
                logger.info(f"Model parameters saved in '{SHEET_LOG_PARAMS}' sheet")
            if LOG_TO_FILE:
                log_path = OUTPUT_DIR / f"QA_{self.timestamp}.log"
                logger.info(f"Text log saved to: {log_path}")
            logger.info("=" * 70)

            return True

        except KeyboardInterrupt:
            logger.warning("Processing interrupted by user (Ctrl+C)")
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Progress saved to: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as e:
                    logger.error(f"Could not save progress: {e}")
            return False

        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)

            # Try to save current state
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Emergency save completed: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as save_error:
                    logger.error(f"Could not save file: {save_error}")

            return False

        finally:
            # Close workbook
            if self.workbook:
                try:
                    self.workbook.close()
                except Exception:
                    pass  # Ignore close errors


def main():
    """Main entry point for the script."""
    try:
        filler = CategoryFiller()
        success = filler.run()

        if success:
            logger.info("Script completed successfully")
        else:
            logger.error("Script completed with errors")

        # Exit with appropriate code
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("\nScript interrupted by user")
        sys.exit(130)  # Standard exit code for Ctrl+C
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
