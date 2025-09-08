#!/usr/bin/env python3
"""DuckDB vector database updater for Q&A system.

This script synchronizes Q&A data from Excel with a DuckDB vector database,
generating embeddings for semantic search capabilities.

The script processes questions and answers from an Excel file, creates
embeddings using GigaChat, and maintains a synchronized vector database
with detailed logging for monitoring and debugging.

Usage:
   python update_db.py

Configuration:
   Adjust settings at the beginning of the script for IDE execution.
"""

import json
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================================
# CONFIGURATION SECTION - Adjust these settings for IDE execution
# ============================================================================

# File paths
INPUT_FILE = Path("in/QA.xlsx")
OUTPUT_DIR = Path("out")
DB_PATH = "qa.duckdb"

# Sheet names
SHEET_QA = "QA"
SHEET_LOG_PREFIX = "LOG_DB"
SHEET_LOG_PARAMS = "LOG_DB_PARAMS"

# Column indices (1-based for openpyxl)
COL_CATEGORY = 1  # Column A in QA sheet
COL_QUESTION = 2  # Column B in QA sheet
COL_ANSWER = 3  # Column C in QA sheet

# Database configuration
EMBEDDING_SIZE = 2560  # GigaChat EmbeddingsGigaR model output size

# Logging configuration
LOG_DB = True  # Set to False to disable database logging sheet
LOG_TO_FILE = True  # Enable logging to text file
LOG_LEVEL = logging.INFO

# Processing configuration
START_ROW = 2  # Skip header row
SAVE_FREQUENCY = 5  # Save file every N processed rows

# Resume configuration
RESUME_FILE = Path(".update_db_resume.json")  # Hidden file for resume state

# ============================================================================
# END OF CONFIGURATION SECTION
# ============================================================================


def setup_logging(log_file: Path | None = None) -> None:
    """Configure unified logging for all modules.

    Sets up consistent logging configuration for the main script and all imported
    modules to ensure uniform formatting and output handling.

    Args:
        log_file: Optional path to log file. If provided, logs will be written
            to both console and file.
    """
    # Create formatter with consistent format
    formatter = logging.Formatter(
        "[%(asctime)s][%(name)s][%(levelname)s][%(filename)s:%(lineno)d][%(message)s]"
    )

    # Get root logger to configure all loggers
    root_logger = logging.getLogger()
    root_logger.setLevel(LOG_LEVEL)

    # Remove any existing handlers to avoid duplicates
    root_logger.handlers.clear()

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(LOG_LEVEL)
    console_handler.setFormatter(formatter)
    root_logger.addHandler(console_handler)

    # File handler if requested
    if log_file and LOG_TO_FILE:
        try:
            log_file.parent.mkdir(parents=True, exist_ok=True)
            file_handler = logging.FileHandler(log_file, encoding="utf-8")
            file_handler.setLevel(LOG_LEVEL)
            file_handler.setFormatter(formatter)
            root_logger.addHandler(file_handler)

            # Log to confirm file logging is active
            logger = logging.getLogger(__name__)
            logger.info(f"Logging to file: {log_file}")
        except Exception as e:
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not create log file: {e}")


# Initialize logging early to capture import messages
setup_logging()

# Import custom modules after logging setup
try:
    from db import duckdb_qa_store
    from embeddings import base_embedding
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Could not import required modules: {e}")
    logger.error("Ensure modules are in the correct paths:")
    logger.error("  - embeddings/base_embedding.py")
    logger.error("  - db/duckdb_qa_store.py")
    sys.exit(1)

# Get logger for this module
logger = logging.getLogger(__name__)


class DatabaseUpdater:
    """Main class for updating Q&A vector database from Excel."""

    def __init__(self):
        """Initialize the database updater."""
        self.workbook: Workbook | None = None
        self.output_file: Path | None = None
        self.db_store: duckdb_qa_store.QADatabaseStore | None = None
        self.processed_count: int = 0
        self.updated_count: int = 0
        self.inserted_count: int = 0
        self.skipped_count: int = 0
        self.resume_state: dict = {}
        self.timestamp: str = ""

    def validate_input_file(self) -> bool:
        """Validate that input file exists and has required sheets.

        Returns:
            True if validation passes, False otherwise.
        """
        if not INPUT_FILE.exists():
            logger.error(f"Input file not found: {INPUT_FILE}")
            return False

        # Check if file is accessible
        try:
            with open(INPUT_FILE, "rb") as f:
                pass
        except PermissionError:
            logger.error(f"Cannot access file: {INPUT_FILE}")
            logger.error("The file might be open in Excel or another program.")
            return False
        except OSError as e:
            logger.error(f"Cannot read file {INPUT_FILE}: {e}")
            return False

        try:
            wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)

            if SHEET_QA not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_QA}' not found")
                wb.close()
                return False

            wb.close()
            return True

        except Exception as e:
            logger.error(f"Error reading input file: {e}")
            return False

    def validate_qa_data(self, sheet: Worksheet) -> bool:
        """Validate QA sheet data completeness.

        Args:
            sheet: The QA worksheet.

        Returns:
            True if all data is complete, False otherwise.
        """
        empty_rows = 0
        incomplete_rows = []

        for row_idx in range(START_ROW, sheet.max_row + 1):
            category = sheet.cell(row=row_idx, column=COL_CATEGORY).value
            question = sheet.cell(row=row_idx, column=COL_QUESTION).value
            answer = sheet.cell(row=row_idx, column=COL_ANSWER).value

            # Skip completely empty rows
            if category is None and question is None and answer is None:
                empty_rows += 1
                continue

            # Check for incomplete data
            if category is None or question is None or answer is None:
                incomplete_rows.append(
                    {
                        "row": row_idx,
                        "category": "Missing" if category is None else "Present",
                        "question": "Missing" if question is None else "Present",
                        "answer": "Missing" if answer is None else "Present",
                    }
                )

        if incomplete_rows:
            logger.error(f"Found {len(incomplete_rows)} incomplete rows:")
            for row_info in incomplete_rows[:5]:
                logger.error(
                    f"  Row {row_info['row']}: "
                    f"Category={row_info['category']}, "
                    f"Question={row_info['question']}, "
                    f"Answer={row_info['answer']}"
                )
            if len(incomplete_rows) > 5:
                logger.error(f"  ... and {len(incomplete_rows) - 5} more")
            return False

        total_rows = sheet.max_row - START_ROW + 1 - empty_rows
        logger.info(f"Validated {total_rows} complete QA rows")
        return True

    def initialize_database(self) -> duckdb_qa_store.QADatabaseStore:
        """Initialize or open the DuckDB database.

        Returns:
            QADatabaseStore instance.
        """
        db_exists = Path(DB_PATH).exists()

        if db_exists:
            logger.info(f"Opening existing database: {DB_PATH}")
        else:
            logger.info(f"Creating new database: {DB_PATH}")

        try:
            db_store = duckdb_qa_store.QADatabaseStore(
                db_path=DB_PATH, embedding_size=EMBEDDING_SIZE
            )

            if db_exists:
                record_count = len(db_store.get_all_qa_records())
                logger.info(f"Database opened successfully with {record_count} records")
            else:
                logger.info("New database created successfully")

            return db_store

        except Exception as e:
            logger.error(f"Failed to initialize database: {e}")
            raise

    def create_log_params_sheet(self, wb: Workbook) -> None:
        """Create or update LOG_DB_PARAMS sheet with script parameters.

        Args:
            wb: Workbook to add the sheet to.
        """
        if not LOG_DB:
            return

        # Remove existing sheet if present
        if SHEET_LOG_PARAMS in wb.sheetnames:
            wb.remove(wb[SHEET_LOG_PARAMS])

        # Create new sheet
        params_sheet = wb.create_sheet(SHEET_LOG_PARAMS)

        # Add headers
        params_sheet.cell(row=1, column=1, value="Parameter")
        params_sheet.cell(row=1, column=2, value="Value")
        params_sheet.cell(row=1, column=3, value="Description")

        # Add parameters with constant names
        row = 2
        config_params = [
            ("Timestamp", self.timestamp, "Processing start time"),
            ("DB_PATH", str(DB_PATH), "DuckDB database path"),
            (
                "DEFAULT_MODEL",
                base_embedding.DEFAULT_MODEL,
                "Default embedding model from base_embedding",
            ),
            ("EMBEDDING_SIZE", EMBEDDING_SIZE, "Embedding vector dimension"),
            ("SAVE_FREQUENCY", SAVE_FREQUENCY, "Save file every N rows"),
            ("START_ROW", START_ROW, "First data row in Excel"),
            ("LOG_TO_FILE", LOG_TO_FILE, "Whether text logging is enabled"),
            ("LOG_LEVEL", logging.getLevelName(LOG_LEVEL), "Logging detail level"),
            ("LOG_DB", LOG_DB, "Whether database logging sheet is enabled"),
            ("INPUT_FILE", str(INPUT_FILE), "Source Excel file path"),
            ("OUTPUT_DIR", str(OUTPUT_DIR), "Output directory for results"),
            ("SHEET_QA", SHEET_QA, "Name of Q&A sheet"),
            ("SHEET_LOG_PREFIX", SHEET_LOG_PREFIX, "Name of logging sheet"),
            ("SHEET_LOG_PARAMS", SHEET_LOG_PARAMS, "Name of parameters sheet"),
            ("COL_CATEGORY", COL_CATEGORY, "Category column index (1-based)"),
            ("COL_QUESTION", COL_QUESTION, "Question column index (1-based)"),
            ("COL_ANSWER", COL_ANSWER, "Answer column index (1-based)"),
            ("RESUME_FILE", str(RESUME_FILE), "Path to resume state file"),
        ]

        for param_name, param_value, param_desc in config_params:
            params_sheet.cell(row=row, column=1, value=param_name)
            params_sheet.cell(row=row, column=2, value=str(param_value))
            params_sheet.cell(row=row, column=3, value=param_desc)
            row += 1

        # Auto-adjust column widths
        for column in params_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            params_sheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Created '{SHEET_LOG_PARAMS}' sheet with script parameters")

    def create_output_file(self) -> tuple[Workbook, Path]:
        """Create output file with timestamp.

        Returns:
            Tuple of (workbook, output_path).
        """
        # Ensure output directory exists
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Generate timestamp
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = OUTPUT_DIR / f"QA_{self.timestamp}.xlsx"

        # Setup file logging for this session
        if LOG_TO_FILE:
            log_file = OUTPUT_DIR / f"QA_{self.timestamp}.log"
            setup_logging(log_file)
        else:
            # Ensure console logging is still active
            setup_logging()

        logger.info(f"Copying {INPUT_FILE} to {output_file}")

        # Copy the file
        shutil.copy2(INPUT_FILE, output_file)
        logger.info("File copied successfully")

        # Open the copied file for modifications
        wb = openpyxl.load_workbook(output_file)

        # Handle LOG_DB sheet based on configuration
        log_sheet_name = SHEET_LOG_PREFIX

        if LOG_DB:
            # Remove existing log sheet if it exists
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing existing '{log_sheet_name}' sheet")
                wb.remove(wb[log_sheet_name])

            # Create new log sheet
            log_sheet = wb.create_sheet(log_sheet_name)

            # Add headers
            headers = ["question", "input_text_q", "answer", "input_text_a"]
            for col_idx, header in enumerate(headers, 1):
                log_sheet.cell(row=1, column=col_idx, value=header)

            logger.info(f"Created '{log_sheet_name}' sheet with headers")

            # Create LOG_DB_PARAMS sheet
            self.create_log_params_sheet(wb)
        else:
            # Remove log sheets if they exist and LOG_DB is False
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing '{log_sheet_name}' sheet (LOG_DB=False)")
                wb.remove(wb[log_sheet_name])
            if SHEET_LOG_PARAMS in wb.sheetnames:
                logger.info(f"Removing '{SHEET_LOG_PARAMS}' sheet (LOG_DB=False)")
                wb.remove(wb[SHEET_LOG_PARAMS])

        # Save modifications
        wb.save(output_file)
        logger.info(f"Created output file: {output_file}")

        return wb, output_file

    def save_resume_state(self, output_file: Path, last_row: int) -> None:
        """Save resume state to file.

        Args:
            output_file: Path to the output file being processed.
            last_row: Last successfully processed row number.
        """
        state = {
            "output_file": str(output_file),
            "last_row": last_row,
            "timestamp": datetime.now().isoformat(),
        }

        try:
            with open(RESUME_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2, ensure_ascii=False)
            logger.debug(f"Saved resume state at row {last_row}")
        except Exception as e:
            logger.warning(f"Could not save resume state: {e}")

    def load_resume_state(self) -> dict | None:
        """Load resume state if exists.

        Returns:
            Resume state dictionary or None if not found.
        """
        if not RESUME_FILE.exists():
            return None

        try:
            with open(RESUME_FILE, encoding="utf-8") as f:
                state: dict = json.load(f)

            # Validate output file still exists
            output_path = Path(state["output_file"])
            if output_path.exists():
                logger.info(f"Found resume state from {state['timestamp']}")
                logger.info(f"Will continue from row {state['last_row'] + 1}")
                return state
            logger.info("Resume output file not found, starting fresh")
            return None

        except Exception as e:
            logger.warning(f"Could not load resume state: {e}")
            return None

    def clear_resume_state(self) -> None:
        """Clear resume state file."""
        if RESUME_FILE.exists():
            try:
                RESUME_FILE.unlink()
                logger.debug("Cleared resume state")
            except Exception as e:
                logger.warning(f"Could not clear resume state: {e}")

    def process_row(
        self, sheet_qa: Worksheet, sheet_log: Worksheet | None, row_idx: int
    ) -> bool:
        """Process a single row from QA sheet.

        Args:
            sheet_qa: The QA worksheet.
            sheet_log: The LOG_DB worksheet (if enabled).
            row_idx: Row index to process.

        Returns:
            True if processing was successful, False otherwise.
        """
        try:
            # Get data from QA sheet
            category = sheet_qa.cell(row=row_idx, column=COL_CATEGORY).value
            question = sheet_qa.cell(row=row_idx, column=COL_QUESTION).value
            answer = sheet_qa.cell(row=row_idx, column=COL_ANSWER).value

            # Convert to strings
            category = str(category).strip() if category else None
            question = str(question).strip()
            answer = str(answer).strip()

            # Log processing start
            question_preview = question[:80] + "..." if len(question) > 80 else question
            logger.info(f"Processing row {row_idx}: {question_preview}")

            # Check if question exists in database
            if self.db_store:
                existing_record = self.db_store.find_question(question)
            else:
                logger.error("Database store not initialized")
                return False

            if existing_record:
                # Question exists - check if answer needs update
                if existing_record["answer"] == answer:
                    logger.info(
                        f"Row {row_idx}: Question exists with same answer, skipping"
                    )
                    self.skipped_count += 1
                    input_text_q = ""
                    input_text_a = ""
                else:
                    logger.info(f"Row {row_idx}: Updating answer for existing question")

                    # Generate new answer embedding
                    logger.info(
                        f"Row {row_idx}: Generating new ANSWER embedding for update..."
                    )
                    answer_emb, answer_inputs = base_embedding.create_embeddings(
                        answer, task_type="document"
                    )

                    # Update in database
                    if self.db_store:
                        success = self.db_store.update_qa(
                            question=question,
                            answer=answer,
                            answer_embedding=answer_emb[0],
                        )
                    else:
                        logger.error("Database store not initialized")
                        return False

                    if success:
                        self.updated_count += 1
                        input_text_q = ""
                        input_text_a = answer_inputs[0]
                        logger.info(f"Row {row_idx}: Successfully updated answer")
                    else:
                        logger.error(f"Row {row_idx}: Failed to update answer")
                        return False
            else:
                # New question - insert with embeddings
                logger.info(f"Row {row_idx}: New question, generating embeddings")

                # Generate embeddings for question
                logger.info(f"Row {row_idx}: Generating QUESTION embedding...")
                question_emb, question_inputs = base_embedding.create_embeddings(
                    question, task_type="query"
                )

                # Generate embeddings for answer
                logger.info(f"Row {row_idx}: Generating ANSWER embedding...")
                answer_emb, answer_inputs = base_embedding.create_embeddings(
                    answer, task_type="document"
                )

                # Insert into database
                if self.db_store:
                    success = self.db_store.insert_qa(
                        question=question,
                        answer=answer,
                        category=category,
                        question_embedding=question_emb[0],
                        answer_embedding=answer_emb[0],
                    )
                else:
                    logger.error("Database store not initialized")
                    return False

                if success:
                    self.inserted_count += 1
                    input_text_q = question_inputs[0]
                    input_text_a = answer_inputs[0]
                    logger.info(f"Row {row_idx}: Successfully inserted new Q&A")
                else:
                    logger.error(f"Row {row_idx}: Failed to insert Q&A")
                    return False

            # Log to LOG_DB sheet if enabled
            if LOG_DB and sheet_log:
                sheet_log.cell(row=row_idx, column=1, value=question)
                sheet_log.cell(row=row_idx, column=2, value=input_text_q)
                sheet_log.cell(row=row_idx, column=3, value=answer)
                sheet_log.cell(row=row_idx, column=4, value=input_text_a)

            self.processed_count += 1
            return True

        except Exception as e:
            logger.error(f"Error processing row {row_idx}: {e}")
            return False

    def delete_obsolete_records(self, sheet: Worksheet) -> int:
        """Delete database records not present in the Excel sheet.

        Args:
            sheet: The QA worksheet.

        Returns:
            Number of deleted records.
        """
        # Collect all current questions from Excel
        current_questions = []

        for row_idx in range(START_ROW, sheet.max_row + 1):
            question = sheet.cell(row=row_idx, column=COL_QUESTION).value
            if question:
                current_questions.append(str(question).strip())

        # Delete missing records
        if self.db_store:
            deleted_count = self.db_store.delete_missing_records(current_questions)
        else:
            logger.error("Database store not initialized")
            return 0

        if deleted_count > 0:
            logger.info(f"Deleted {deleted_count} obsolete records from database")

        return deleted_count

    def run(self) -> bool:
        """Main execution method.

        Returns:
            True if execution completed successfully, False otherwise.
        """
        try:
            logger.info("=" * 70)
            logger.info("Starting Database Update Script")
            logger.info("=" * 70)

            # Check for resume state
            resume_state = self.load_resume_state()

            if resume_state:
                # Resume from previous run
                logger.info("Resuming from previous run...")
                self.output_file = Path(resume_state["output_file"])

                # Extract timestamp from filename for log file
                filename = self.output_file.stem
                if filename.startswith("QA_"):
                    self.timestamp = filename[3:]

                # Setup file logging for resumed session
                if LOG_TO_FILE:
                    log_file = OUTPUT_DIR / f"QA_{self.timestamp}.log"
                    setup_logging(log_file)
                    logger.info("=" * 70)
                    logger.info("RESUMED SESSION")
                    logger.info("=" * 70)

                self.workbook = openpyxl.load_workbook(self.output_file)
                start_from_row = resume_state["last_row"] + 1
            else:
                # Fresh start
                logger.info("Starting fresh processing...")

                # Step 1: Validate input file
                logger.info("Step 1: Validating input file...")
                if not self.validate_input_file():
                    logger.error("FAILED: Cannot proceed with processing")
                    return False

                # Step 2: Create output file (this also sets up file logging)
                logger.info("Step 2: Creating output file...")
                self.workbook, self.output_file = self.create_output_file()
                start_from_row = START_ROW

            # Step 3: Validate QA data completeness
            logger.info("Step 3: Validating QA sheet data...")
            sheet_qa = self.workbook[SHEET_QA]

            if not self.validate_qa_data(sheet_qa):
                logger.error(
                    "Data validation failed. Please ensure all columns A, B, C are filled"
                )
                return False

            # Step 4: Initialize database
            logger.info("Step 4: Initializing database...")
            self.db_store = self.initialize_database()

            # Get LOG_DB sheet if enabled
            sheet_log = None
            if LOG_DB:
                if SHEET_LOG_PREFIX in self.workbook.sheetnames:
                    sheet_log = self.workbook[SHEET_LOG_PREFIX]
                else:
                    logger.warning(
                        f"LOG_DB is enabled but '{SHEET_LOG_PREFIX}' sheet not found"
                    )

            # Step 5: Process rows
            logger.info("Step 5: Synchronizing data with database...")
            logger.info("-" * 70)

            # Get total rows to process
            total_rows = sheet_qa.max_row - START_ROW + 1

            for row_idx in range(start_from_row, sheet_qa.max_row + 1):
                # Check if row has data
                question = sheet_qa.cell(row=row_idx, column=COL_QUESTION).value
                if not question:
                    continue

                # Process row
                success = self.process_row(sheet_qa, sheet_log, row_idx)

                if not success:
                    logger.error(f"Failed to process row {row_idx}")
                    # Save progress before stopping
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx - 1)
                        logger.info(
                            f"Progress saved. You can resume from row {row_idx}"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save progress: {save_error}")
                    return False

                # Periodic save
                if (row_idx - start_from_row + 1) % SAVE_FREQUENCY == 0:
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        progress = row_idx - START_ROW + 1
                        logger.info(
                            f"Progress saved: {progress}/{total_rows} rows processed"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save progress: {save_error}")

            # Step 6: Delete obsolete records
            logger.info("Step 6: Cleaning up obsolete records...")
            deleted_count = self.delete_obsolete_records(sheet_qa)

            # Final save
            self.workbook.save(self.output_file)
            logger.info(f"Final save completed: {self.output_file}")

            # Clear resume state on successful completion
            self.clear_resume_state()

            # Summary
            logger.info("-" * 70)
            logger.info("Database update completed successfully!")
            logger.info(f"Total rows processed: {self.processed_count}")
            logger.info(f"  - Inserted: {self.inserted_count}")
            logger.info(f"  - Updated: {self.updated_count}")
            logger.info(f"  - Skipped: {self.skipped_count}")
            logger.info(f"  - Deleted from DB: {deleted_count}")
            logger.info(f"Output file: {self.output_file}")
            if LOG_DB and sheet_log:
                logger.info(f"Detailed logs saved in '{SHEET_LOG_PREFIX}' sheet")
                logger.info(f"Script parameters saved in '{SHEET_LOG_PARAMS}' sheet")
            if LOG_TO_FILE:
                log_path = OUTPUT_DIR / f"QA_{self.timestamp}.log"
                logger.info(f"Text log saved to: {log_path}")
            logger.info("=" * 70)

            return True

        except KeyboardInterrupt:
            logger.warning("Processing interrupted by user (Ctrl+C)")
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Progress saved to: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as e:
                    logger.error(f"Could not save progress: {e}")
            return False

        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)

            # Try to save current state
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Emergency save completed: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as save_error:
                    logger.error(f"Could not save file: {save_error}")

            return False

        finally:
            # Clean up resources
            if self.db_store:
                try:
                    self.db_store.close()
                except Exception:
                    pass

            if self.workbook:
                try:
                    self.workbook.close()
                except Exception:
                    pass


def main():
    """Main entry point for the script."""
    try:
        updater = DatabaseUpdater()
        success = updater.run()

        if success:
            logger.info("Script completed successfully")
        else:
            logger.error("Script completed with errors")

        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("\nScript interrupted by user")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
