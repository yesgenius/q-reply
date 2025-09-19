#!/usr/bin/env python3
"""Conference answer judgement script using symmetric entailment evaluation.

This script evaluates the quality of generated answers by comparing them with
reference answers using bidirectional semantic entailment. It processes candidate
answers from a QT file, compares them with reference answers from QA file, and
generates comprehensive evaluation metrics with detailed logging support.

The script calculates precision, recall, F1 scores, detects contradictions and
hallucinations, and provides quality classification (good/ok/bad) for each answer.

Usage:
    python judge_answer.py

Configuration:
    Adjust settings at the beginning of the script for IDE execution.
"""

from __future__ import annotations

from datetime import datetime
import json
import logging
from pathlib import Path
import shutil
import statistics
import sys
import time
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Import centralized logging
from utils.logger import close_logging, get_logger, setup_logging


# ============================================================================
# CONFIGURATION SECTION - Adjust these settings for IDE execution
# ============================================================================

# File paths
INPUT_FILE_QT = Path("in/QT.xlsx")  # Copy of Q.xlsx with answers
INPUT_FILE_QA = Path("in/QA.xlsx")  # Reference Q&A for comparison
OUTPUT_DIR = Path("out")

# Sheet names
SHEET_Q = "Q"
SHEET_QA = "QA"
SHEET_LOG_PREFIX = "LOG_JUDGEMENT"
SHEET_LOG_PARAMS = "LOG_JUDGEMENT_PARAMS"

# Column indices for Q sheet (1-based for openpyxl)
COL_Q_QUESTION = 1  # Column A - Question
COL_Q_ANSWER = 2  # Column B - Answer

# Column indices for QA sheet (1-based for openpyxl)
COL_QA_QUESTION = 2  # Column B - Question
COL_QA_ANSWER = 3  # Column C - Answer

# Retry configuration
MAX_RETRY_ATTEMPTS_JUDGEMENT = 3  # Maximum retry attempts for judgement
RETRY_DELAY = 2  # Delay between retries in seconds

# Logging configuration
LOG_JUDGEMENT = True  # Set to False to disable judgement logging sheet
LOG_TO_FILE = True  # Enable logging to text file
LOG_LEVEL = logging.INFO

# Processing configuration
START_ROW = 2  # Skip header row
SAVE_FREQUENCY = 3  # Save file every N processed rows

# Resume configuration
RESUME_FILE = Path(".judge_answer_resume.json")  # Hidden file for resume state

# ============================================================================
# END OF CONFIGURATION SECTION
# ============================================================================

# Initialize module logger
logger = get_logger(__name__)

# Import required modules after logging setup
try:
    from prompts import get_judgement_prompt
except ImportError as e:
    logger.error(f"Error: Could not import required modules: {e}")
    logger.error("Ensure all modules are in the correct paths:")
    logger.error("  - prompts/get_judgement_prompt.py")
    sys.exit(1)


def format_json_for_excel(data: Any) -> str:
    """Format JSON data for Excel cell display.

    Formats JSON data (messages or raw responses) for readable display
    in Excel cells while preserving the complete data structure.

    Args:
        data: JSON-serializable data (dict, list, or already stringified JSON).

    Returns:
        Formatted JSON string suitable for Excel cell display.
    """
    try:
        # If already a string, try to parse and re-format
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                # Already formatted or not JSON, return as-is
                return str(data)

        # Format with indentation and ensure_ascii=False for readability
        formatted = json.dumps(data, ensure_ascii=False, indent=2)

        # Replace escaped newlines with actual newlines for Excel
        formatted = formatted.replace("\\n", "\n")

        return formatted

    except Exception as e:
        logger.warning(f"Could not format JSON for Excel: {e}")
        # Return string representation as fallback
        return str(data)


class JudgeAnswer:
    """Main class for evaluating answers using symmetric entailment judgement."""

    def __init__(self):
        """Initialize the answer judge."""
        self.workbook: Workbook | None = None
        self.output_file: Path | None = None
        self.processed_count: int = 0
        self.resume_state: dict = {}
        self.timestamp: str = ""
        self.reference_data: dict[int, dict[str, str]] = {}  # Changed to dict with row index
        self.first_result_column: int = 0  # Store column for results
        self.qa_max_row: int = 0  # Store max row from QA file
        # Statistics tracking
        self.scores: list[float] = []
        self.classes: dict[str, int] = {"good": 0, "ok": 0, "bad": 0}
        self.contradiction_count: int = 0
        self.hallucination_count: int = 0

    def extract_response_content(self, raw_response: Any) -> str:
        """Extract content field from LLM response.

        Args:
            raw_response: Raw response from LLM (dict or str).

        Returns:
            Content field value or empty string on error.
        """
        try:
            # Convert to dict if string
            if isinstance(raw_response, str):
                raw_response = json.loads(raw_response)

            # Navigate: choices[0].message.content
            content = raw_response.get("choices", [{}])[0].get("message", {}).get("content", "")
            # Ensure string return type
            return str(content) if content is not None else ""
        except (json.JSONDecodeError, KeyError, IndexError, TypeError, AttributeError):
            return ""

    def validate_input_files(self) -> bool:
        """Validate that required input files exist and are accessible.

        Returns:
            True if validation passes, False otherwise.
        """
        # Check QT.xlsx
        if not INPUT_FILE_QT.exists():
            logger.error(f"Input file not found: {INPUT_FILE_QT}")
            return False

        # Check if QT.xlsx is accessible
        try:
            with open(INPUT_FILE_QT, "rb") as f:
                pass
        except PermissionError:
            logger.error(f"Cannot access file: {INPUT_FILE_QT}")
            logger.error("The file might be open in Excel or another program.")
            return False
        except OSError as e:
            logger.error(f"Cannot read file {INPUT_FILE_QT}: {e}")
            return False

        # Check QT.xlsx structure
        try:
            wb = openpyxl.load_workbook(INPUT_FILE_QT, read_only=True)
            if SHEET_Q not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_Q}' not found in {INPUT_FILE_QT}")
                wb.close()
                return False
            wb.close()
        except Exception as e:
            logger.error(f"Error reading {INPUT_FILE_QT}: {e}")
            return False

        # Check QA.xlsx
        if not INPUT_FILE_QA.exists():
            logger.error(f"Reference file not found: {INPUT_FILE_QA}")
            return False

        # Check QA.xlsx structure
        try:
            wb = openpyxl.load_workbook(INPUT_FILE_QA, read_only=True)
            if SHEET_QA not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_QA}' not found in {INPUT_FILE_QA}")
                wb.close()
                return False
            wb.close()
        except Exception as e:
            logger.error(f"Error reading {INPUT_FILE_QA}: {e}")
            return False

        return True

    def load_reference_data(self) -> bool:
        """Load reference Q&A data from QA.xlsx preserving row alignment.

        Returns:
            True if data loaded successfully, False otherwise.
        """
        try:
            wb = openpyxl.load_workbook(INPUT_FILE_QA, read_only=True)
            sheet_qa = wb[SHEET_QA]

            # Store max row for later use
            self.qa_max_row = sheet_qa.max_row

            self.reference_data = {}
            empty_count = 0

            # Load all rows, preserving positions for row-by-row correspondence
            for row_idx in range(START_ROW, sheet_qa.max_row + 1):
                question = sheet_qa.cell(row=row_idx, column=COL_QA_QUESTION).value
                answer = sheet_qa.cell(row=row_idx, column=COL_QA_ANSWER).value

                # Store data even if empty to preserve row alignment
                if question is None and answer is None:
                    empty_count += 1
                    # Store empty reference to maintain alignment
                    self.reference_data[row_idx] = {"question": "", "answer": ""}
                elif question is None or answer is None:
                    logger.warning(f"Incomplete reference data at row {row_idx}")
                    # Store partial data to maintain alignment
                    self.reference_data[row_idx] = {
                        "question": str(question).strip() if question else "",
                        "answer": str(answer).strip() if answer else "",
                    }
                else:
                    self.reference_data[row_idx] = {
                        "question": str(question).strip(),
                        "answer": str(answer).strip(),
                    }

            wb.close()

            valid_count = len(
                [r for r in self.reference_data.values() if r["question"] and r["answer"]]
            )

            if valid_count == 0:
                logger.error("No valid reference data found in QA.xlsx")
                return False

            logger.info(
                f"Loaded {len(self.reference_data)} reference rows ({valid_count} valid, {empty_count} empty)"
            )
            return True

        except Exception as e:
            logger.error(f"Error loading reference data: {e}")
            return False

    def create_log_params_sheet(self, wb: Workbook) -> None:
        """Create or update LOG_JUDGEMENT_PARAMS sheet with model parameters.

        Args:
            wb: Workbook to add the sheet to.
        """
        if not LOG_JUDGEMENT:
            return

        # Remove existing sheet if present
        if SHEET_LOG_PARAMS in wb.sheetnames:
            wb.remove(wb[SHEET_LOG_PARAMS])

        # Create new sheet
        params_sheet = wb.create_sheet(SHEET_LOG_PARAMS)

        # Add headers
        params_sheet.cell(row=1, column=1, value="Parameter")
        params_sheet.cell(row=1, column=2, value="Value")
        params_sheet.cell(row=1, column=3, value="Description")

        # Configuration parameters - including file paths and column indices
        row = 2
        config_params = [
            ("Timestamp", self.timestamp, "Processing start time"),
            # File paths
            ("INPUT_FILE_QT", str(INPUT_FILE_QT), "Input file with candidate answers"),
            ("INPUT_FILE_QA", str(INPUT_FILE_QA), "Reference Q&A file"),
            ("OUTPUT_DIR", str(OUTPUT_DIR), "Output directory"),
            # Sheet configuration
            ("SHEET_Q", SHEET_Q, "Question sheet name"),
            ("SHEET_QA", SHEET_QA, "Reference QA sheet name"),
            # Column configuration
            ("COL_Q_QUESTION", COL_Q_QUESTION, "Question column in Q sheet"),
            ("COL_Q_ANSWER", COL_Q_ANSWER, "Answer column in Q sheet"),
            ("COL_QA_QUESTION", COL_QA_QUESTION, "Question column in QA sheet"),
            ("COL_QA_ANSWER", COL_QA_ANSWER, "Answer column in QA sheet"),
            # Processing parameters
            (
                "MAX_RETRY_ATTEMPTS_JUDGEMENT",
                MAX_RETRY_ATTEMPTS_JUDGEMENT,
                "Max retries for judgement",
            ),
            ("RETRY_DELAY", RETRY_DELAY, "Delay between retries (seconds)"),
            ("SAVE_FREQUENCY", SAVE_FREQUENCY, "Save file every N rows"),
            ("LOG_JUDGEMENT", LOG_JUDGEMENT, "Enable detailed logging sheet"),
            ("LOG_TO_FILE", LOG_TO_FILE, "Enable logging to text file"),
            ("LOG_LEVEL", logging.getLevelName(LOG_LEVEL), "Logging detail level"),
            ("START_ROW", START_ROW, "First row to process (skip headers)"),
        ]

        for param_name, param_value, param_desc in config_params:
            params_sheet.cell(row=row, column=1, value=param_name)
            params_sheet.cell(row=row, column=2, value=str(param_value))
            params_sheet.cell(row=row, column=3, value=param_desc)
            row += 1

        # Add separator for judgement model parameters
        try:
            judgement_params = getattr(get_judgement_prompt, "params", {})
            if judgement_params:
                row += 1
                params_sheet.cell(row=row, column=1, value="--- Judgement Model Parameters ---")
                row += 1

                param_descriptions = {
                    "model": "LLM model name for judgement",
                    "temperature": "Randomness of responses (0-1)",
                    "top_p": "Nucleus sampling threshold",
                    "stream": "Whether to stream responses",
                    "max_tokens": "Maximum tokens in response",
                }

                for param_name, param_value in judgement_params.items():
                    params_sheet.cell(row=row, column=1, value=f"judgement.{param_name}")
                    params_sheet.cell(row=row, column=2, value=str(param_value))
                    params_sheet.cell(
                        row=row, column=3, value=param_descriptions.get(param_name, "")
                    )
                    row += 1
        except AttributeError:
            logger.debug("Judgement prompt module params not available")

        # Add evaluation configuration constants
        row += 1
        params_sheet.cell(row=row, column=1, value="--- Evaluation Configuration ---")
        row += 1

        eval_params = [
            ("SCORE_SCALE", 100, "Score scale (0-100)"),
            ("THRESHOLD_GOOD", 85, "Threshold for 'good' classification"),
            ("THRESHOLD_OK", 70, "Threshold for 'ok' classification"),
            ("CONTRADICTION_PENALTY", 0.20, "Penalty for contradiction"),
            ("HALLUCINATION_PENALTY", 0.10, "Penalty for hallucination"),
            ("NUMERICAL_TOLERANCE_ABSOLUTE", 1e-6, "Absolute numerical tolerance"),
            ("NUMERICAL_TOLERANCE_RELATIVE", 0.02, "Relative numerical tolerance (2%)"),
        ]

        for param_name, param_value, param_desc in eval_params:
            params_sheet.cell(row=row, column=1, value=param_name)
            params_sheet.cell(row=row, column=2, value=str(param_value))
            params_sheet.cell(row=row, column=3, value=param_desc)
            row += 1

        # Auto-adjust column widths
        for column in params_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            params_sheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Created '{SHEET_LOG_PARAMS}' sheet with model parameters")

    def create_output_file(self) -> tuple[Workbook, Path]:
        """Create output file with timestamp.

        Returns:
            Tuple of (workbook, output_path).
        """
        # Ensure output directory exists
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Generate timestamp
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = OUTPUT_DIR / f"QT_{self.timestamp}.xlsx"

        # Setup file logging for this session
        if LOG_TO_FILE:
            log_file = OUTPUT_DIR / f"QT_{self.timestamp}.log"
            setup_logging(log_file=log_file, level=LOG_LEVEL)

        logger.info(f"Copying {INPUT_FILE_QT} to {output_file}")

        # Copy file to preserve structure
        shutil.copy2(INPUT_FILE_QT, output_file)
        logger.info("File copied successfully")

        # Open the copied file for modifications
        wb = openpyxl.load_workbook(output_file)

        # Get Q sheet and find first free column
        sheet_q = wb[SHEET_Q]
        self.first_result_column = sheet_q.max_column + 1  # Store for later use

        # Add headers for judgement results
        judgement_headers = [
            "reference_question",
            "reference_answer",
            "score",
            "class",
            "f1",
            "precision_c_to_r",
            "recall_r_to_c",
            "contradiction",
            "hallucination",
            "justification",
            "evidence",
            "penalties",
        ]

        for col_idx, header in enumerate(judgement_headers):
            sheet_q.cell(row=1, column=self.first_result_column + col_idx, value=header)

        logger.info(f"Added judgement headers starting from column {self.first_result_column}")

        # Handle LOG_JUDGEMENT sheet based on configuration
        log_sheet_name = SHEET_LOG_PREFIX

        if LOG_JUDGEMENT:
            # Remove existing LOG_JUDGEMENT sheet if it exists
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing existing '{log_sheet_name}' sheet for fresh start")
                wb.remove(wb[log_sheet_name])

            # Create new LOG_JUDGEMENT sheet
            log_sheet = wb.create_sheet(log_sheet_name)

            # Add headers to LOG_JUDGEMENT sheet
            headers = [
                "candidate_question",
                "candidate_answer",
                "reference_question",
                "reference_answer",
                "score",
                "class",
                "f1",
                "precision_c_to_r",
                "recall_r_to_c",
                "contradiction",
                "hallucination",
                "justification",
                "evidence",
                "penalties",
                "messages",
                "response",
                "response_content",
            ]
            for col_idx, header in enumerate(headers, 1):
                log_sheet.cell(row=1, column=col_idx, value=header)

            logger.info(f"Created '{log_sheet_name}' sheet with headers")

            # Create LOG_JUDGEMENT_PARAMS sheet
            self.create_log_params_sheet(wb)
        else:
            # If LOG_JUDGEMENT is False and sheet exists, remove it
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing '{log_sheet_name}' sheet (LOG_JUDGEMENT=False)")
                wb.remove(wb[log_sheet_name])
            if SHEET_LOG_PARAMS in wb.sheetnames:
                logger.info(f"Removing '{SHEET_LOG_PARAMS}' sheet (LOG_JUDGEMENT=False)")
                wb.remove(wb[SHEET_LOG_PARAMS])

        # Save modifications
        wb.save(output_file)
        logger.info(f"Created output file: {output_file}")

        return wb, output_file

    def save_resume_state(self, output_file: Path, last_row: int) -> None:
        """Save resume state to file.

        Args:
            output_file: Path to the output file being processed.
            last_row: Last successfully processed row number.
        """
        state = {
            "output_file": str(output_file),
            "last_row": last_row,
            "timestamp": datetime.now().isoformat(),
        }

        try:
            with open(RESUME_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2, ensure_ascii=False)
            logger.debug(f"Saved resume state at row {last_row}")
        except Exception as e:
            logger.warning(f"Could not save resume state: {e}")

    def load_resume_state(self) -> dict[Any, Any] | None:
        """Load resume state if exists.

        Returns:
            Resume state dictionary or None if not found.
        """
        if not RESUME_FILE.exists():
            return None

        try:
            with open(RESUME_FILE, encoding="utf-8") as f:
                state: dict[Any, Any] = json.load(f)

            # Validate output file still exists
            output_path = Path(state["output_file"])
            if output_path.exists():
                logger.info(f"Found resume state from {state['timestamp']}")
                logger.info(f"Will continue from row {state['last_row'] + 1}")
                return state
            logger.warning("Resume output file not found, starting fresh")
            return None

        except Exception as e:
            logger.warning(f"Could not load resume state: {e}")
            return None

    def clear_resume_state(self) -> None:
        """Clear resume state file."""
        if RESUME_FILE.exists():
            try:
                RESUME_FILE.unlink()
                logger.debug("Cleared resume state")
            except Exception as e:
                logger.warning(f"Could not clear resume state: {e}")

    def judge_answer_with_retry(
        self, question: str, reference_answer: str, candidate_answer: str
    ) -> dict[str, Any]:
        """Judge answer with retry logic on errors.

        Args:
            question: The original question.
            reference_answer: Reference answer from QA.
            candidate_answer: Candidate answer from QT.

        Returns:
            Judgement result dictionary.
        """
        last_error = None

        for attempt in range(1, MAX_RETRY_ATTEMPTS_JUDGEMENT + 1):
            try:
                # get_judgement_prompt.run returns 3 values
                result_json, messages_list, raw_response = get_judgement_prompt.run(
                    question=question,
                    reference_answer=reference_answer,
                    candidate_answer=candidate_answer,
                )
                result: dict[str, Any] = json.loads(result_json)

                # Format messages for logging
                result["messages"] = format_json_for_excel(messages_list)

                # Format raw response for logging
                result["response"] = format_json_for_excel(raw_response)

                # Extract content from response
                result["response_content"] = self.extract_response_content(raw_response)

                if attempt > 1:
                    logger.info(f"Judgement succeeded on attempt {attempt}")

                return result

            except Exception as e:
                last_error = e
                logger.warning(
                    f"Judgement attempt {attempt}/{MAX_RETRY_ATTEMPTS_JUDGEMENT} failed: {e}"
                )

                if attempt < MAX_RETRY_ATTEMPTS_JUDGEMENT:
                    if RETRY_DELAY > 0:
                        time.sleep(RETRY_DELAY)
                    continue

        # All attempts failed - return error result
        logger.error(f"All {MAX_RETRY_ATTEMPTS_JUDGEMENT} judgement attempts failed")
        return {
            "score": 0,
            "class": "bad",
            "f1": 0.0,
            "precision_c_to_r": 0.0,
            "recall_r_to_c": 0.0,
            "contradiction": False,
            "hallucination": False,
            "justification": f"Failed after {MAX_RETRY_ATTEMPTS_JUDGEMENT} attempts. Error: {last_error!s}",
            "evidence": [],
            "penalties": 0.0,
            "messages": "[]",
            "response": "",
            "response_content": "",
        }

    def process_row(self, sheet_q: Worksheet, sheet_log: Worksheet | None, row_idx: int) -> bool:
        """Process a single row from Q sheet.

        Args:
            sheet_q: The Q worksheet.
            sheet_log: The LOG_JUDGEMENT worksheet (if enabled).
            row_idx: Row index to process.

        Returns:
            True if processing was successful, False otherwise.
        """
        try:
            # Get candidate data from QT
            candidate_question = sheet_q.cell(row=row_idx, column=COL_Q_QUESTION).value
            candidate_answer = sheet_q.cell(row=row_idx, column=COL_Q_ANSWER).value

            if candidate_question is None:
                logger.warning(f"Empty question at row {row_idx}, skipping")
                return True

            candidate_question_str = str(candidate_question).strip()
            candidate_answer_str = str(candidate_answer).strip() if candidate_answer else ""

            # Get corresponding reference data using direct row mapping
            reference = self.reference_data.get(row_idx, {"question": "", "answer": ""})
            reference_question = reference["question"]
            reference_answer = reference["answer"]

            # Skip if no valid reference data
            if not reference_answer:
                logger.warning(f"No reference answer for row {row_idx}, skipping")
                return True

            question_preview = (
                candidate_question_str[:80] + "..."
                if len(candidate_question_str) > 80
                else candidate_question_str
            )
            logger.info(f"Processing row {row_idx}: {question_preview}")

            # Judge the answer with retry
            judgement_result = self.judge_answer_with_retry(
                candidate_question_str, reference_answer, candidate_answer_str
            )

            # Update statistics
            score = judgement_result.get("score", 0)
            self.scores.append(score)

            quality_class = judgement_result.get("class", "bad")
            self.classes[quality_class] = self.classes.get(quality_class, 0) + 1

            if judgement_result.get("contradiction", False):
                self.contradiction_count += 1
            if judgement_result.get("hallucination", False):
                self.hallucination_count += 1

            # Write results to Q sheet using stored column position
            col = self.first_result_column
            sheet_q.cell(row=row_idx, column=col, value=reference_question)
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=reference_answer)
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=judgement_result.get("score", 0))
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=judgement_result.get("class", ""))
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=judgement_result.get("f1", 0.0))
            col += 1
            sheet_q.cell(
                row=row_idx, column=col, value=judgement_result.get("precision_c_to_r", 0.0)
            )
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=judgement_result.get("recall_r_to_c", 0.0))
            col += 1
            sheet_q.cell(
                row=row_idx, column=col, value=judgement_result.get("contradiction", False)
            )
            col += 1
            sheet_q.cell(
                row=row_idx, column=col, value=judgement_result.get("hallucination", False)
            )
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=judgement_result.get("justification", ""))
            col += 1
            # Format evidence as string for Excel
            evidence = judgement_result.get("evidence", [])
            evidence_str = json.dumps(evidence, ensure_ascii=False)
            sheet_q.cell(row=row_idx, column=col, value=evidence_str)
            col += 1
            sheet_q.cell(row=row_idx, column=col, value=judgement_result.get("penalties", 0.0))

            # Write to LOG_JUDGEMENT sheet if enabled
            if LOG_JUDGEMENT and sheet_log:
                col = 1
                sheet_log.cell(row=row_idx, column=col, value=candidate_question_str)
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=candidate_answer_str)
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=reference_question)
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=reference_answer)
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=judgement_result.get("score", 0))
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=judgement_result.get("class", ""))
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=judgement_result.get("f1", 0.0))
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("precision_c_to_r", 0.0)
                )
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("recall_r_to_c", 0.0)
                )
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("contradiction", False)
                )
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("hallucination", False)
                )
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("justification", "")
                )
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=evidence_str)
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("penalties", 0.0)
                )
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=judgement_result.get("messages", ""))
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=judgement_result.get("response", ""))
                col += 1
                sheet_log.cell(
                    row=row_idx, column=col, value=judgement_result.get("response_content", "")
                )

            logger.info(f"Row {row_idx} processed: score={score}, class={quality_class}")
            self.processed_count += 1
            return True

        except Exception as e:
            logger.error(f"Error processing row {row_idx}: {e}")
            return False

    def calculate_aggregates(self) -> dict[str, Any]:
        """Calculate aggregate statistics from processed results.

        Returns:
            Dictionary with aggregate metrics.
        """
        if not self.scores:
            return {
                "mean_score": 0.0,
                "median_score": 0.0,
                "stdev_score": 0.0,
                "share_good": 0.0,
                "share_ok": 0.0,
                "share_bad": 0.0,
                "contradiction_rate": 0.0,
                "hallucination_rate": 0.0,
                "total_processed": 0,
            }

        total = len(self.scores)

        # Calculate score statistics
        mean_score = statistics.mean(self.scores)
        median_score = statistics.median(self.scores)
        stdev_score = statistics.stdev(self.scores) if len(self.scores) > 1 else 0.0

        # Calculate class shares
        share_good = self.classes.get("good", 0) / total
        share_ok = self.classes.get("ok", 0) / total
        share_bad = self.classes.get("bad", 0) / total

        # Calculate flag rates
        contradiction_rate = self.contradiction_count / total
        hallucination_rate = self.hallucination_count / total

        return {
            "mean_score": round(mean_score, 2),
            "median_score": round(median_score, 2),
            "stdev_score": round(stdev_score, 2),
            "share_good": round(share_good, 3),
            "share_ok": round(share_ok, 3),
            "share_bad": round(share_bad, 3),
            "contradiction_rate": round(contradiction_rate, 3),
            "hallucination_rate": round(hallucination_rate, 3),
            "total_processed": total,
        }

    def run(self) -> bool:
        """Main execution method.

        Returns:
            True if execution completed successfully, False otherwise.
        """
        try:
            logger.info("=" * 70)
            logger.info("Starting Judge Answer Script")
            logger.info(f"Log Judgement: {'ENABLED' if LOG_JUDGEMENT else 'DISABLED'}")
            logger.info(f"Max Retry Attempts: {MAX_RETRY_ATTEMPTS_JUDGEMENT}")
            logger.info("=" * 70)

            # Check for resume state
            resume_state = self.load_resume_state()

            if resume_state:
                # Resume from previous run
                logger.info("Resuming from previous run...")
                self.output_file = Path(resume_state["output_file"])

                # Extract timestamp from filename
                filename = self.output_file.stem  # QT_YYYY-MM-DD_HHMMSS
                if filename.startswith("QT_"):
                    self.timestamp = filename[3:]  # Remove "QT_" prefix

                # Setup file logging for resumed session
                if LOG_TO_FILE:
                    log_file = OUTPUT_DIR / f"QT_{self.timestamp}.log"
                    setup_logging(log_file=log_file, level=LOG_LEVEL)
                    logger.info("=" * 70)
                    logger.info("RESUMED SESSION")
                    logger.info("=" * 70)

                self.workbook = openpyxl.load_workbook(self.output_file)

                # Load reference data for resumed session
                logger.info("Loading reference data for resumed session...")
                if not self.load_reference_data():
                    logger.error("Failed to load reference data")
                    return False

                # Restore first_result_column from existing headers
                sheet_q = self.workbook[SHEET_Q]
                for col in range(1, sheet_q.max_column + 1):
                    if sheet_q.cell(row=1, column=col).value == "reference_question":
                        self.first_result_column = col
                        logger.info(f"Restored result column position: {self.first_result_column}")
                        break

                start_from_row = resume_state["last_row"] + 1
            else:
                # Fresh start
                logger.info("Starting fresh processing...")

                # Step 1: Validate input files
                logger.info("Step 1: Validating input files...")
                if not self.validate_input_files():
                    logger.error("=" * 70)
                    logger.error("FAILED: Cannot proceed with processing")
                    logger.error("=" * 70)
                    return False

                # Step 2: Load reference data
                logger.info("Step 2: Loading reference data...")
                if not self.load_reference_data():
                    logger.error("Failed to load reference data")
                    return False

                # Step 3: Create output file
                logger.info("Step 3: Creating output file...")
                self.workbook, self.output_file = self.create_output_file()

                start_from_row = START_ROW

            # Step 4: Initialize judgement system
            logger.info("Step 4: Initializing judgement system...")
            get_judgement_prompt.update_system_prompt()
            logger.info("Judgement system initialized")

            # Step 5: Process answers
            logger.info("Step 5: Processing answers...")

            sheet_q = self.workbook[SHEET_Q]

            # Get LOG_JUDGEMENT sheet if enabled
            sheet_log = None
            if LOG_JUDGEMENT:
                if SHEET_LOG_PREFIX in self.workbook.sheetnames:
                    sheet_log = self.workbook[SHEET_LOG_PREFIX]
                else:
                    logger.warning(
                        f"LOG_JUDGEMENT is enabled but '{SHEET_LOG_PREFIX}' sheet not found"
                    )

            # Determine rows to process based on QA max row
            rows_to_process = []
            for row_idx in range(start_from_row, min(sheet_q.max_row + 1, self.qa_max_row + 1)):
                question = sheet_q.cell(row=row_idx, column=COL_Q_QUESTION).value
                if question and str(question).strip():
                    rows_to_process.append(row_idx)

            if not rows_to_process:
                logger.info("No questions to process")
                self.clear_resume_state()
                return True

            logger.info(f"Found {len(rows_to_process)} answers to judge")
            logger.info("-" * 70)

            # Process each row
            successfully_processed: list[int] = []

            for idx, row_idx in enumerate(rows_to_process, 1):
                # Process row
                success = self.process_row(sheet_q, sheet_log, row_idx)

                if not success:
                    logger.error(f"Failed to process row {row_idx}")
                    # Save progress on error
                    if successfully_processed:
                        try:
                            self.workbook.save(self.output_file)
                            last_successful = successfully_processed[-1]
                            self.save_resume_state(self.output_file, last_successful)
                            logger.info(
                                f"Emergency save after error: saved up to row {last_successful}"
                            )
                        except Exception as save_error:
                            logger.error(f"Could not perform emergency save: {save_error}")
                    continue

                # Track successful processing
                successfully_processed.append(row_idx)

                # Periodic save
                if idx % SAVE_FREQUENCY == 0:
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(f"Progress saved: {idx}/{len(rows_to_process)} rows processed")
                    except Exception as save_error:
                        logger.error(f"Could not save progress: {save_error}")

                # Final save
                if idx == len(rows_to_process):
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(
                            f"Final batch saved: {idx}/{len(rows_to_process)} rows completed"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save final batch: {save_error}")

            # Final save
            self.workbook.save(self.output_file)
            logger.info(f"Final save completed: {self.output_file}")

            # Clear resume state on successful completion
            self.clear_resume_state()

            # Calculate aggregates
            aggregates = self.calculate_aggregates()

            # Summary
            logger.info("-" * 70)
            logger.info("Processing completed successfully!")
            logger.info(f"Total rows processed: {self.processed_count}")
            logger.info(f"Output file: {self.output_file}")

            # Display aggregates
            logger.info("-" * 70)
            logger.info("AGGREGATE STATISTICS:")
            logger.info(f"Mean Score: {aggregates['mean_score']}")
            logger.info(f"Median Score: {aggregates['median_score']}")
            logger.info(f"StDev Score: {aggregates['stdev_score']}")
            logger.info(f"Share Good: {aggregates['share_good']:.1%}")
            logger.info(f"Share OK: {aggregates['share_ok']:.1%}")
            logger.info(f"Share Bad: {aggregates['share_bad']:.1%}")
            logger.info(f"Contradiction Rate: {aggregates['contradiction_rate']:.1%}")
            logger.info(f"Hallucination Rate: {aggregates['hallucination_rate']:.1%}")

            if LOG_JUDGEMENT and sheet_log:
                logger.info(f"Detailed logs saved in '{SHEET_LOG_PREFIX}' sheet")
                logger.info(f"Model parameters saved in '{SHEET_LOG_PARAMS}' sheet")
            if LOG_TO_FILE:
                log_path = OUTPUT_DIR / f"QT_{self.timestamp}.log"
                logger.info(f"Text log saved to: {log_path}")
            logger.info("=" * 70)

            return True

        except KeyboardInterrupt:
            logger.warning("Processing interrupted by user (Ctrl+C)")
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Progress saved to: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as e:
                    logger.error(f"Could not save progress: {e}")
            return False

        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)

            # Try to save current state
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Emergency save completed: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as save_error:
                    logger.error(f"Could not save file: {save_error}")

            return False

        finally:
            # Close workbook
            if self.workbook:
                try:
                    self.workbook.close()
                except Exception:
                    pass


def main():
    """Main entry point for the script."""
    try:
        # Initialize logging for the main application
        setup_logging(level=LOG_LEVEL)

        judge = JudgeAnswer()
        success = judge.run()

        if success:
            logger.info("Script completed successfully")
        else:
            logger.error("Script completed with errors")

        # Exit with appropriate code
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("\nScript interrupted by user")
        sys.exit(130)  # Standard exit code for Ctrl+C
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)
    finally:
        # Clean up logging resources
        close_logging()


if __name__ == "__main__":
    main()
