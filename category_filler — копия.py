#!/usr/bin/env python3
"""Question category filler script with answer context support.

This script automatically fills question categories using GigaChat LLM
based on a predefined category dictionary from an Excel file.
Now supports using answers to enhance categorization accuracy.

The script processes questions from an Excel file, categorizes them using
the define_category_prompt module with optional answer context, and saves
results with detailed logging.

Usage:
    python category_filler.py

Configuration:
    Adjust settings at the beginning of the script for IDE execution.
"""


import json
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Import the categorization module
try:
    from prompts import define_category_prompt
except ImportError:
    print("Error: Could not import define_category_prompt module")
    print("Ensure the module is in the correct path: prompts/define_category_prompt.py")
    sys.exit(1)

# ============================================================================
# CONFIGURATION SECTION - Adjust these settings for IDE execution
# ============================================================================

# File paths
INPUT_FILE = Path("in/QA.xlsx")
OUTPUT_DIR = Path("out")

# Sheet names
SHEET_QA = "QA"
SHEET_CATEGORY = "CATEGORY"
SHEET_LOG_PREFIX = "LOG_CATEGORY"

# Column indices (1-based for openpyxl)
COL_CATEGORY_NAME = 1  # Column A in CATEGORY sheet
COL_CATEGORY_DESC = 2  # Column B in CATEGORY sheet
COL_QA_CATEGORY = 1  # Column A in QA sheet
COL_QA_QUESTION = 2  # Column B in QA sheet
COL_QA_ANSWER = 3  # Column C in QA sheet

# Categorization configuration
USE_ANSWER_FOR_CATEGORIZATION = (
    True  # Include answer in categorization for better accuracy
)

# Logging configuration
LOG_CATEGORY = True  # Set to False to disable category logging sheet
LOG_LEVEL = logging.INFO

# Processing configuration
START_ROW = 2  # Skip header row
SAVE_FREQUENCY = 5  # Save file every N processed rows (reduced for safety)

# Resume configuration
RESUME_FILE = Path(".category_filler_resume.json")  # Hidden file for resume state

# ============================================================================
# END OF CONFIGURATION SECTION
# ============================================================================

# Configure logging
logging.basicConfig(
    level=LOG_LEVEL,
    format="[%(asctime)s][%(name)s][%(levelname)s][%(filename)s:%(lineno)d][%(message)s]",
)
logger = logging.getLogger(__name__)


class CategoryFiller:
    """Main class for filling question categories using LLM with answer context support."""

    def __init__(self):
        """Initialize the category filler."""
        self.workbook: Optional[Workbook] = None
        self.output_file: Optional[Path] = None
        self.categories: Dict[str, str] = {}
        self.processed_count: int = 0
        self.resume_state: Dict = {}
        self.timestamp: str = ""

    def validate_input_file(self) -> bool:
        """Validate that input file exists and has required sheets.

        Also checks if the file is accessible (not locked by another program).

        Returns:
            True if validation passes, False otherwise.
        """
        if not INPUT_FILE.exists():
            logger.error(f"Input file not found: {INPUT_FILE}")
            return False

        # Check if file is accessible (not locked)
        try:
            # Try to open file in exclusive mode to check if it's locked
            with open(INPUT_FILE, "rb") as f:
                # Just open and close - if this works, file is not locked
                pass
        except PermissionError:
            logger.error(f"Cannot access file: {INPUT_FILE}")
            logger.error("The file might be open in Excel or another program.")
            logger.error("Please close the file and run the script again.")
            return False
        except IOError as e:
            logger.error(f"Cannot read file {INPUT_FILE}: {e}")
            return False

        try:
            wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)

            # Check required sheets
            if SHEET_CATEGORY not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_CATEGORY}' not found")
                wb.close()
                return False

            if SHEET_QA not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_QA}' not found")
                wb.close()
                return False

            wb.close()
            return True

        except PermissionError:
            logger.error(f"File is locked: {INPUT_FILE}")
            logger.error(
                "Please close the file in Excel or other programs and try again."
            )
            return False
        except Exception as e:
            logger.error(f"Error reading input file: {e}")
            return False

    def load_categories(self, sheet: Worksheet) -> Dict[str, str]:
        """Load category dictionary from CATEGORY sheet.

        Args:
            sheet: The CATEGORY worksheet.

        Returns:
            Dictionary mapping category names to descriptions.

        Raises:
            ValueError: If categories have missing data.
        """
        categories = {}
        empty_rows = 0

        for row_idx in range(START_ROW, sheet.max_row + 1):
            category_name = sheet.cell(row=row_idx, column=COL_CATEGORY_NAME).value
            category_desc = sheet.cell(row=row_idx, column=COL_CATEGORY_DESC).value

            # Skip completely empty rows
            if category_name is None and category_desc is None:
                empty_rows += 1
                continue

            # Validate completeness - both fields must be filled
            if category_name is None or category_desc is None:
                raise ValueError(
                    f"Incomplete category data at row {row_idx} in CATEGORY sheet. "
                    f"Category name: {category_name}, Description: {category_desc}. "
                    f"Both columns A and B must be filled."
                )

            categories[str(category_name).strip()] = str(category_desc).strip()

        if not categories:
            raise ValueError(
                "No categories found in CATEGORY sheet. Please fill columns A and B."
            )

        logger.info(f"Loaded {len(categories)} categories from CATEGORY sheet")
        return categories

    def validate_qa_data(self, sheet: Worksheet) -> List[int]:
        """Validate QA sheet data completeness and identify rows needing processing.

        Args:
            sheet: The QA worksheet.

        Returns:
            List of row indices that need category processing.

        Raises:
            ValueError: If questions or answers are missing.
        """
        rows_to_process = []
        total_rows = 0

        for row_idx in range(START_ROW, sheet.max_row + 1):
            question = sheet.cell(row=row_idx, column=COL_QA_QUESTION).value
            answer = sheet.cell(row=row_idx, column=COL_QA_ANSWER).value
            category = sheet.cell(row=row_idx, column=COL_QA_CATEGORY).value

            # Skip completely empty rows
            if question is None and answer is None and category is None:
                continue

            total_rows += 1

            # Validate that both question and answer are present
            if question is None or answer is None:
                raise ValueError(
                    f"Incomplete QA data at row {row_idx} in QA sheet. "
                    f"Question (column B): {'Present' if question else 'Missing'}, "
                    f"Answer (column C): {'Present' if answer else 'Missing'}. "
                    f"Both columns B and C must be filled."
                )

            # Handle multi-line categories - treat as string and check each line
            if category is not None:
                category_str = str(category).strip()
                # Check if it's a multi-line category
                if "\n" in category_str:
                    # Split and check if ALL lines are valid categories
                    category_lines = [
                        line.strip()
                        for line in category_str.split("\n")
                        if line.strip()
                    ]
                    all_valid = all(cat in self.categories for cat in category_lines)
                    if not all_valid:
                        rows_to_process.append(row_idx)
                        logger.debug(
                            f"Row {row_idx}: Multi-line category contains invalid entries, needs processing"
                        )
                elif category_str == "":
                    rows_to_process.append(row_idx)
                    logger.debug(f"Row {row_idx}: Category is empty, needs processing")
                elif category_str not in self.categories:
                    rows_to_process.append(row_idx)
                    logger.debug(
                        f"Row {row_idx}: Category '{category_str}' not in dictionary, needs processing"
                    )
            else:
                rows_to_process.append(row_idx)
                logger.debug(f"Row {row_idx}: Category is None, needs processing")

        logger.info(f"Validated {total_rows} QA rows")
        return rows_to_process

    def create_output_file(self) -> Tuple[Workbook, Path]:
        """Create output file with timestamp.

        Returns:
            Tuple of (workbook, output_path).
        """
        # Ensure output directory exists
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Generate timestamp
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = OUTPUT_DIR / f"QA_{self.timestamp}.xlsx"

        logger.info(f"Copying {INPUT_FILE} to {output_file}")

        # SIMPLE FILE COPY - preserves EVERYTHING exactly as is
        shutil.copy2(INPUT_FILE, output_file)
        logger.info(f"File copied successfully")

        # Now open the copied file for modifications
        wb = openpyxl.load_workbook(output_file)

        # Handle LOG_CATEGORY sheet based on configuration
        log_sheet_name = SHEET_LOG_PREFIX

        if LOG_CATEGORY:
            # Remove existing LOG_CATEGORY sheet if it exists (for fresh start)
            if log_sheet_name in wb.sheetnames:
                logger.info(
                    f"Removing existing '{log_sheet_name}' sheet for fresh start"
                )
                wb.remove(wb[log_sheet_name])

            # Create new LOG_CATEGORY sheet
            log_sheet = wb.create_sheet(log_sheet_name)

            # Add headers to LOG_CATEGORY sheet
            headers = [
                "Категории",
                "Вопросы",
                "category",
                "confidence",
                "reasoning",
                "messages",
            ]
            for col_idx, header in enumerate(headers, 1):
                log_sheet.cell(row=1, column=col_idx, value=header)

            logger.info(f"Created '{log_sheet_name}' sheet with headers")
        else:
            # If LOG_CATEGORY is False and sheet exists, remove it
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing '{log_sheet_name}' sheet (LOG_CATEGORY=False)")
                wb.remove(wb[log_sheet_name])

        # Save modifications (only LOG_CATEGORY changes, everything else intact)
        wb.save(output_file)
        logger.info(f"Created output file: {output_file}")

        return wb, output_file

    def save_resume_state(self, output_file: Path, last_row: int) -> None:
        """Save resume state to file.

        Args:
            output_file: Path to the output file being processed.
            last_row: Last successfully processed row number.
        """
        state = {
            "output_file": str(output_file),
            "last_row": last_row,
            "timestamp": datetime.now().isoformat(),
        }

        try:
            with open(RESUME_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2, ensure_ascii=False)
            logger.debug(f"Saved resume state at row {last_row}")
        except Exception as e:
            logger.warning(f"Could not save resume state: {e}")

    def load_resume_state(self) -> Optional[Dict]:
        """Load resume state if exists.

        Returns:
            Resume state dictionary or None if not found.
        """
        if not RESUME_FILE.exists():
            return None

        try:
            with open(RESUME_FILE, "r", encoding="utf-8") as f:
                state = json.load(f)

            # Validate output file still exists
            output_path = Path(state["output_file"])
            if output_path.exists():
                logger.info(f"Found resume state from {state['timestamp']}")
                logger.info(f"Will continue from row {state['last_row'] + 1}")
                return state
            else:
                logger.warning("Resume output file not found, starting fresh")
                return None

        except Exception as e:
            logger.warning(f"Could not load resume state: {e}")
            return None

    def clear_resume_state(self) -> None:
        """Clear resume state file."""
        if RESUME_FILE.exists():
            try:
                RESUME_FILE.unlink()
                logger.debug("Cleared resume state")
            except Exception as e:
                logger.warning(f"Could not clear resume state: {e}")

    def format_json_for_excel(self, json_string: str) -> str:
        """Format JSON string for better readability in Excel cells.

        Replaces escaped newlines with actual line breaks for Excel display.

        Args:
            json_string: JSON string potentially containing escaped newlines.

        Returns:
            Formatted string with actual line breaks for Excel.
        """
        # Replace escaped newlines with actual line breaks
        # This makes multi-line content readable in Excel cells
        return json_string.replace("\\n", "\n")

    def categorize_question(
        self, question: str, answer: Optional[str] = None
    ) -> Dict[str, Any]:
        """Categorize a single question using LLM with optional answer context.

        Args:
            question: The question text to categorize.
            answer: Optional answer text to provide additional context.

        Returns:
            Dictionary with category, confidence, reasoning, and messages.
        """
        try:
            # Run categorization - answer will be used if provided and USE_ANSWER_FOR_CATEGORIZATION is True
            result_json, messages_list = define_category_prompt.run(
                question,
                answer=answer if USE_ANSWER_FOR_CATEGORIZATION and answer else None,
            )
            result = json.loads(result_json)

            # Convert messages list to JSON string for storage
            messages_json = json.dumps(messages_list, ensure_ascii=False, indent=2)

            # Format JSON for Excel readability
            messages_formatted = messages_json.replace("\\n", "\n")

            # Add formatted messages to result dictionary
            result["messages"] = messages_formatted

            return result

        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON response: {e}")
            return {
                "category": "Error",
                "confidence": 0.0,
                "reasoning": f"JSON parse error: {str(e)}",
                "messages": "[]",
            }
        except Exception as e:
            logger.error(f"Categorization failed: {e}")
            return {
                "category": "Error",
                "confidence": 0.0,
                "reasoning": f"Categorization error: {str(e)}",
                "messages": "[]",
            }

    def process_row(
        self, sheet_qa: Worksheet, sheet_log: Optional[Worksheet], row_idx: int
    ) -> bool:
        """Process a single row from QA sheet.

        Args:
            sheet_qa: The QA worksheet.
            sheet_log: The LOG_CATEGORY worksheet (if enabled).
            row_idx: Row index to process.

        Returns:
            True if processing was successful, False otherwise.
        """
        try:
            # Get data from QA sheet
            current_category = sheet_qa.cell(row=row_idx, column=COL_QA_CATEGORY).value
            question = sheet_qa.cell(row=row_idx, column=COL_QA_QUESTION).value
            answer = sheet_qa.cell(row=row_idx, column=COL_QA_ANSWER).value

            if question is None:
                logger.warning(f"Empty question at row {row_idx}, skipping")
                return True

            # Log processing start
            question_preview = (
                str(question)[:80] + "..." if len(str(question)) > 80 else str(question)
            )
            logger.info(f"Processing row {row_idx}: {question_preview}")

            # Log if using answer context
            if USE_ANSWER_FOR_CATEGORIZATION and answer:
                answer_preview = (
                    str(answer)[:50] + "..." if len(str(answer)) > 50 else str(answer)
                )
                logger.debug(f"Using answer for categorization: {answer_preview}")

            # Categorize the question with optional answer
            result = self.categorize_question(
                str(question), str(answer) if answer else None
            )

            # Update category in QA sheet
            sheet_qa.cell(row=row_idx, column=COL_QA_CATEGORY, value=result["category"])

            # Log to LOG_CATEGORY sheet if enabled
            if LOG_CATEGORY and sheet_log:
                # Write to the same row number as in QA sheet for alignment
                sheet_log.cell(row=row_idx, column=1, value=current_category or "")
                sheet_log.cell(row=row_idx, column=2, value=question)
                sheet_log.cell(row=row_idx, column=3, value=result["category"])
                sheet_log.cell(row=row_idx, column=4, value=result["confidence"])
                sheet_log.cell(row=row_idx, column=5, value=result.get("reasoning", ""))
                sheet_log.cell(row=row_idx, column=6, value=result.get("messages", ""))

            # Log result
            logger.info(
                f"Row {row_idx} processed: '{result['category']}' "
                f"(confidence: {result['confidence']:.2f})"
            )

            self.processed_count += 1
            return True

        except Exception as e:
            logger.error(f"Error processing row {row_idx}: {e}")
            return False

    def run(self) -> bool:
        """Main execution method.

        Returns:
            True if execution completed successfully, False otherwise.
        """
        try:
            logger.info("=" * 70)
            logger.info("Starting Category Filler Script")
            logger.info(
                f"Answer Usage: {'ENABLED' if USE_ANSWER_FOR_CATEGORIZATION else 'DISABLED'}"
            )
            logger.info("=" * 70)

            # Check for resume state
            resume_state = self.load_resume_state()

            if resume_state:
                # Resume from previous run
                logger.info("Resuming from previous run...")
                self.output_file = Path(resume_state["output_file"])
                self.workbook = openpyxl.load_workbook(self.output_file)
                start_from_row = resume_state["last_row"] + 1

                # Extract timestamp from filename for consistency
                filename = self.output_file.stem  # QA_YYYY-MM-DD_HHMMSS
                if filename.startswith("QA_"):
                    self.timestamp = filename[3:]  # Remove "QA_" prefix

                # Load categories from existing file
                if SHEET_CATEGORY not in self.workbook.sheetnames:
                    logger.error(f"Sheet '{SHEET_CATEGORY}' not found in resume file")
                    return False

                sheet_category = self.workbook[SHEET_CATEGORY]
                self.categories = self.load_categories(sheet_category)

            else:
                # Fresh start
                logger.info("Starting fresh processing...")

                # Step 1: Validate input file
                logger.info("Step 1: Validating input file...")
                if not self.validate_input_file():
                    logger.error("=" * 70)
                    logger.error("FAILED: Cannot proceed with processing")
                    logger.error("=" * 70)
                    return False

                # Step 2: Create output file
                logger.info("Step 2: Creating output file...")
                self.workbook, self.output_file = self.create_output_file()

                # Step 3: Load and validate categories
                logger.info("Step 3: Loading categories from CATEGORY sheet...")
                sheet_category = self.workbook[SHEET_CATEGORY]
                try:
                    self.categories = self.load_categories(sheet_category)
                except ValueError as e:
                    logger.error(f"Category validation failed: {e}")
                    logger.error(
                        "Please ensure all categories in column A have descriptions in column B"
                    )
                    return False

                start_from_row = START_ROW

            # Display loaded categories
            logger.info(f"Categories loaded: {', '.join(self.categories.keys())}")

            # Step 4: Initialize categorization system
            logger.info("Step 4: Initializing GigaChat categorization system...")

            # Initialize with categories
            define_category_prompt.update_system_prompt(categories=self.categories)
            logger.info("Categorization system initialized successfully")

            if USE_ANSWER_FOR_CATEGORIZATION:
                logger.info("Note: Answers will be used for enhanced categorization")

            # Step 5: Validate QA data
            logger.info("Step 5: Validating QA sheet data...")
            sheet_qa = self.workbook[SHEET_QA]

            try:
                rows_to_process = self.validate_qa_data(sheet_qa)
            except ValueError as e:
                logger.error(f"QA data validation failed: {e}")
                logger.error(
                    "Please ensure all questions (column B) and answers (column C) are filled"
                )
                return False

            # Filter rows based on resume state
            if resume_state:
                original_count = len(rows_to_process)
                rows_to_process = [r for r in rows_to_process if r >= start_from_row]
                logger.info(
                    f"Resuming: skipping {original_count - len(rows_to_process)} already processed rows"
                )

            if not rows_to_process:
                logger.info(
                    "No rows need processing. All categories are already filled correctly."
                )
                self.clear_resume_state()
                return True

            logger.info(
                f"Found {len(rows_to_process)} rows that need category assignment"
            )

            # Get LOG_CATEGORY sheet if enabled
            sheet_log = None
            if LOG_CATEGORY:
                if SHEET_LOG_PREFIX in self.workbook.sheetnames:
                    sheet_log = self.workbook[SHEET_LOG_PREFIX]
                else:
                    logger.warning(
                        f"LOG_CATEGORY is enabled but '{SHEET_LOG_PREFIX}' sheet not found"
                    )

            # Step 6: Process rows
            logger.info("Step 6: Processing rows...")
            if USE_ANSWER_FOR_CATEGORIZATION:
                logger.info("Note: Using answers for enhanced categorization accuracy")
            logger.info("-" * 70)

            # Type annotation for mypy
            successfully_processed: List[int] = []  # Track successfully processed rows

            for idx, row_idx in enumerate(rows_to_process, 1):
                # Process row
                success = self.process_row(sheet_qa, sheet_log, row_idx)

                if not success:
                    logger.error(f"Failed to process row {row_idx}")

                    # Always save what we have processed so far (no need for flag)
                    if successfully_processed:
                        try:
                            self.workbook.save(self.output_file)
                            last_successful = successfully_processed[-1]
                            self.save_resume_state(self.output_file, last_successful)
                            logger.info(
                                f"Emergency save after error: saved up to row {last_successful}"
                            )
                        except Exception as save_error:
                            logger.error(
                                f"Could not perform emergency save: {save_error}"
                            )

                    continue  # Skip to next row

                # Track successful processing
                successfully_processed.append(row_idx)

                # Periodic save based on SAVE_FREQUENCY
                if idx % SAVE_FREQUENCY == 0:
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(
                            f"Progress saved: {idx}/{len(rows_to_process)} rows processed (saved up to row {row_idx})"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save progress: {save_error}")
                        # Continue processing even if save fails

                # Also save on the last row
                if idx == len(rows_to_process):
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(
                            f"Final batch saved: {idx}/{len(rows_to_process)} rows completed"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save final batch: {save_error}")

            # Final save
            self.workbook.save(self.output_file)
            logger.info(f"Final save completed: {self.output_file}")

            # Clear resume state on successful completion
            self.clear_resume_state()

            # Summary
            logger.info("-" * 70)
            logger.info("Processing completed successfully!")
            logger.info(f"Total rows processed: {self.processed_count}")
            logger.info(
                f"Answer usage was: {'ENABLED' if USE_ANSWER_FOR_CATEGORIZATION else 'DISABLED'}"
            )
            logger.info(f"Output file: {self.output_file}")
            if LOG_CATEGORY and sheet_log:
                logger.info(f"Detailed logs saved in '{SHEET_LOG_PREFIX}' sheet")
            logger.info("=" * 70)

            return True

        except KeyboardInterrupt:
            logger.warning("Processing interrupted by user (Ctrl+C)")
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Progress saved to: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as e:
                    logger.error(f"Could not save progress: {e}")
            return False

        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)

            # Try to save current state
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Emergency save completed: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as save_error:
                    logger.error(f"Could not save file: {save_error}")

            return False

        finally:
            # Close workbook
            if self.workbook:
                try:
                    self.workbook.close()
                except Exception:
                    pass  # Ignore close errors


def main():
    """Main entry point for the script."""
    try:
        filler = CategoryFiller()
        success = filler.run()

        if success:
            logger.info("Script completed successfully")
        else:
            logger.error("Script completed with errors")

        # Exit with appropriate code
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("\nScript interrupted by user")
        sys.exit(130)  # Standard exit code for Ctrl+C
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
