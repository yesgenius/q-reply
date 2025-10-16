#!/usr/bin/env python3
"""DuckDB to Excel exporter for Q&A system.

This script exports Q&A data from DuckDB vector database to Excel file,
preserving the original database record order.

The script reads all records from the vector database and generates
an Excel file with proper formatting and metadata sheets.

Usage:
    python db_export.py

Configuration:
    Adjust settings at the beginning of the script for IDE execution.
"""

from __future__ import annotations

from datetime import datetime
import logging
from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Import centralized logging
from utils.logger import close_logging, get_logger, setup_logging


# ============================================================================
# CONFIGURATION SECTION - Adjust these settings for IDE execution
# ============================================================================

# File paths
OUTPUT_DIR = Path("out")
DB_PATH = "storages/qa.duckdb"

# Sheet names
SHEET_QA = "QA"
SHEET_EXPORT_INFO = "EXPORT_INFO"

# Column indices (1-based for openpyxl)
COL_CATEGORY = 1  # Column A in QA sheet
COL_QUESTION = 2  # Column B in QA sheet
COL_ANSWER = 3  # Column C in QA sheet

# Logging configuration
LOG_TO_FILE = True  # Enable logging to text file
LOG_LEVEL = logging.INFO

# Export configuration
START_ROW = 2  # Skip header row
INCLUDE_METADATA = True  # Include export metadata sheet

# ============================================================================
# END OF CONFIGURATION SECTION
# ============================================================================

# Initialize module logger
logger = get_logger(__name__)

# Import custom modules after logging setup
try:
    from storages import duckdb_qa_store
except ImportError as e:
    logger.error(f"Could not import required modules: {e}")
    logger.error("Ensure module is in the correct path:")
    logger.error("  - storages/duckdb_qa_store.py")
    sys.exit(1)


class DatabaseExporter:
    """Main class for exporting Q&A data from DuckDB to Excel."""

    def __init__(self):
        """Initialize the database exporter."""
        self.workbook: Workbook | None = None
        self.output_file: Path | None = None
        self.db_store: duckdb_qa_store.QADatabaseStore | None = None
        self.exported_count: int = 0
        self.timestamp: str = ""

    def validate_database(self) -> bool:
        """Validate that database exists and is accessible.

        Returns:
            True if validation passes, False otherwise.
        """
        db_path = Path(DB_PATH)

        if not db_path.exists():
            logger.error(f"Database not found: {DB_PATH}")
            return False

        # Check if file is accessible
        try:
            with open(db_path, "rb") as f:
                pass
        except PermissionError:
            logger.error(f"Cannot access database: {DB_PATH}")
            logger.error("The database might be locked by another process.")
            return False
        except OSError as e:
            logger.error(f"Cannot read database {DB_PATH}: {e}")
            return False

        return True

    def open_database(self) -> duckdb_qa_store.QADatabaseStore:
        """Open the DuckDB database for reading.

        Returns:
            QADatabaseStore instance.

        Raises:
            RuntimeError: If database cannot be opened.
        """
        logger.info(f"Opening database: {DB_PATH}")

        try:
            db_store = duckdb_qa_store.QADatabaseStore(
                db_path=DB_PATH,
                embedding_size=2560,  # Must match the database configuration
            )

            record_count = len(db_store.get_all_qa_records())
            logger.info(f"Database opened successfully with {record_count} records")

            if record_count == 0:
                logger.warning("Database is empty, no records to export")

            return db_store

        except Exception as e:
            logger.error(f"Failed to open database: {e}")
            raise RuntimeError(f"Cannot open database: {e}")

    def create_output_file(self) -> tuple[Workbook, Path]:
        """Create output Excel file with timestamp.

        Returns:
            Tuple of (workbook, output_path).
        """
        # Ensure output directory exists
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Generate timestamp
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = OUTPUT_DIR / f"QA_EXPORT_{self.timestamp}.xlsx"

        # Setup file logging for this session
        if LOG_TO_FILE:
            log_file = OUTPUT_DIR / f"QA_EXPORT_{self.timestamp}.log"
            setup_logging(log_file=log_file, level=LOG_LEVEL)
            logger.info("=" * 70)
            logger.info("DATABASE EXPORT SESSION")
            logger.info("=" * 70)

        logger.info(f"Creating output file: {output_file}")

        # Create new workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create QA sheet
        qa_sheet = wb.create_sheet(SHEET_QA)

        # Add headers
        qa_sheet.cell(row=1, column=COL_CATEGORY, value="Category")
        qa_sheet.cell(row=1, column=COL_QUESTION, value="Question")
        qa_sheet.cell(row=1, column=COL_ANSWER, value="Answer")

        logger.info(f"Created '{SHEET_QA}' sheet with headers")

        return wb, output_file

    def create_metadata_sheet(self, wb: Workbook, record_count: int) -> None:
        """Create metadata sheet with export information.

        Args:
            wb: Workbook to add the sheet to.
            record_count: Number of exported records.
        """
        if not INCLUDE_METADATA:
            return

        # Create metadata sheet
        info_sheet = wb.create_sheet(SHEET_EXPORT_INFO)

        # Add headers
        info_sheet.cell(row=1, column=1, value="Parameter")
        info_sheet.cell(row=1, column=2, value="Value")

        # Add export information
        row = 2
        export_info = [
            ("Export Timestamp", self.timestamp),
            ("Database Path", str(DB_PATH)),
            ("Total Records", record_count),
            ("Output Directory", str(OUTPUT_DIR)),
            ("Sheet Name", SHEET_QA),
            ("Start Row", START_ROW),
        ]

        for param_name, param_value in export_info:
            info_sheet.cell(row=row, column=1, value=param_name)
            info_sheet.cell(row=row, column=2, value=str(param_value))
            row += 1

        # Auto-adjust column widths
        for column in info_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            info_sheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Created '{SHEET_EXPORT_INFO}' sheet with export metadata")

    def export_records(self, qa_sheet: Worksheet) -> int:
        """Export all records from database to Excel sheet.

        Records are exported in the original database order.

        Args:
            qa_sheet: The QA worksheet to write data to.

        Returns:
            Number of exported records.
        """
        if not self.db_store:
            raise RuntimeError("Database store not initialized")

        # Get all records from database
        logger.info("Fetching all records from database...")
        records = self.db_store.get_all_qa_records()

        if not records:
            logger.warning("No records found in database")
            return 0

        logger.info(f"Found {len(records)} records to export")
        logger.info("-" * 70)

        # Write records to Excel in original database order
        row_idx = START_ROW
        for i, record in enumerate(records, 1):
            # Extract fields
            category = record.get("category", "")
            question = record["question"]
            answer = record["answer"]

            # Write to sheet
            qa_sheet.cell(row=row_idx, column=COL_CATEGORY, value=category)
            qa_sheet.cell(row=row_idx, column=COL_QUESTION, value=question)
            qa_sheet.cell(row=row_idx, column=COL_ANSWER, value=answer)

            # Log progress
            if i % 100 == 0 or i == len(records):
                logger.info(f"Exported {i}/{len(records)} records")

            row_idx += 1
            self.exported_count += 1

        # Auto-adjust column widths for better readability
        for column in qa_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        # Limit width calculation to first 100 chars for performance
                        cell_length = min(len(str(cell.value)), 100)
                        max_length = max(max_length, cell_length)
                except:
                    pass

            # Set reasonable width limits
            adjusted_width = max(10, min(max_length + 2, 80))
            qa_sheet.column_dimensions[column_letter].width = adjusted_width

        return self.exported_count

    def run(self) -> bool:
        """Main execution method.

        Returns:
            True if execution completed successfully, False otherwise.
        """
        try:
            logger.info("=" * 70)
            logger.info("Starting Database Export Script")
            logger.info("=" * 70)

            # Step 1: Validate database
            logger.info("Step 1: Validating database...")
            if not self.validate_database():
                logger.error("FAILED: Cannot proceed with export")
                return False

            # Step 2: Open database
            logger.info("Step 2: Opening database...")
            self.db_store = self.open_database()

            # Step 3: Create output file
            logger.info("Step 3: Creating output file...")
            self.workbook, self.output_file = self.create_output_file()

            # Step 4: Export records
            logger.info("Step 4: Exporting records...")
            qa_sheet = self.workbook[SHEET_QA]
            exported_count = self.export_records(qa_sheet)

            # Step 5: Add metadata sheet
            if INCLUDE_METADATA:
                logger.info("Step 5: Adding metadata...")
                self.create_metadata_sheet(self.workbook, exported_count)

            # Step 6: Save file
            logger.info("Step 6: Saving file...")
            self.workbook.save(self.output_file)
            logger.info(f"File saved: {self.output_file}")

            # Summary
            logger.info("-" * 70)
            logger.info("Database export completed successfully!")
            logger.info(f"Total records exported: {self.exported_count}")
            logger.info("Records order preserved from database")
            logger.info(f"Output file: {self.output_file}")
            if INCLUDE_METADATA:
                logger.info(f"Export metadata saved in '{SHEET_EXPORT_INFO}' sheet")
            if LOG_TO_FILE:
                log_path = OUTPUT_DIR / f"QA_EXPORT_{self.timestamp}.log"
                logger.info(f"Export log saved to: {log_path}")
            logger.info("=" * 70)

            return True

        except KeyboardInterrupt:
            logger.warning("Export interrupted by user (Ctrl+C)")
            return False

        except Exception as e:
            logger.error(f"Unexpected error during export: {e}", exc_info=True)
            return False

        finally:
            # Clean up resources
            if self.db_store:
                try:
                    self.db_store.close()
                    logger.debug("Database connection closed")
                except Exception:
                    pass

            if self.workbook:
                try:
                    self.workbook.close()
                    logger.debug("Workbook closed")
                except Exception:
                    pass


def main():
    """Main entry point for the script."""
    try:
        # Initialize logging for the main application
        setup_logging(level=LOG_LEVEL)

        exporter = DatabaseExporter()
        success = exporter.run()

        if success:
            logger.info("Export completed successfully")
        else:
            logger.error("Export completed with errors")

        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("\nScript interrupted by user")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)
    finally:
        # Clean up logging resources
        close_logging()


if __name__ == "__main__":
    main()
