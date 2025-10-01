#!/usr/bin/env python3
"""Conference question answer generator script using RAG with DuckDB.

This script generates answers for conference questions using a Retrieval-Augmented
Generation (RAG) approach. It searches for similar questions in a DuckDB database,
retrieves relevant Q&A pairs as context, and generates comprehensive answers using
the GigaChat LLM.

The script processes questions from an Excel file, performs semantic search using
embeddings, optionally categorizes questions, and generates context-aware answers
with detailed logging support.

Usage:
    python answer_generator.py

Configuration:
    Adjust settings at the beginning of the script for IDE execution.
"""

from __future__ import annotations

from datetime import datetime
from enum import Enum
import json
import logging
from pathlib import Path
import shutil
import sys
import time
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Import centralized logging
from utils.logger import close_logging, get_logger, setup_logging


# ============================================================================
# CONFIGURATION SECTION - Adjust these settings for IDE execution
# ============================================================================

# File paths
INPUT_FILE_Q = Path("in/Q.xlsx")
INPUT_FILE_QA = Path("in/QA.xlsx")
DATABASE_FILE = Path("storages/qa.duckdb")
OUTPUT_DIR = Path("out")

# Sheet names
SHEET_Q = "Q"
SHEET_T = "T"
SHEET_CATEGORY = "CATEGORY"
SHEET_LOG_PREFIX = "LOG_ANSWER"
SHEET_LOG_PARAMS = "LOG_ANSWER_PARAMS"  # New sheet for model parameters

# Column indices for Q sheet (1-based for openpyxl)
COL_Q_QUESTION = 1  # Column A - Question
COL_Q_ANSWER = 2  # Column B - Answer
COL_Q_CONFIDENCE = 3  # Column C - Answer confidence
COL_Q_SOURCES = 4  # Column D - Sources used
COL_Q_S1 = 5  # Column E - First similarity score
COL_Q_Q1 = 6  # Column F - First similar question
COL_Q_A1 = 7  # Column G - First similar answer
# Additional columns for s2/q2/a2, s3/q3/a3, etc. will be dynamic


class CategorySearchMode(Enum):
    """Defines how categories are used in similarity search."""

    WITH_CATEGORY = "with_category"  # Search using detected category
    WITHOUT_CATEGORY = "without_category"  # Ignore categories in search
    CATEGORY_FALLBACK = "category_fallback"  # Try with category first, fallback to without


# Column indices for CATEGORY sheet (1-based for openpyxl)
COL_CATEGORY_NAME = 1  # Column A
COL_CATEGORY_DESC = 2  # Column B

# Search configuration
TOP_K_SIMILAR = 3  # Number of similar questions to retrieve
SIMILARITY_THRESHOLD = 0.8  # Minimum cosine similarity
CATEGORY_SEARCH_MODE = CategorySearchMode.WITHOUT_CATEGORY  # How to use categories in search

# Embedding configuration
EMBEDDING_MODEL = "EmbeddingsGigaR"
EMBEDDING_DIMENSION = 2560

# Retry configuration
MAX_RETRY_ATTEMPTS_CATEGORY = 3  # Maximum retry attempts for categorization
MAX_RETRY_ATTEMPTS_ANSWER = 3  # Maximum retry attempts for answer generation
RETRY_DELAY = 2  # Delay between retries in seconds

# Logging configuration
LOG_ANSWER = True  # Set to False to disable answer logging sheet
LOG_TO_FILE = True  # Enable logging to text file
LOG_LEVEL = logging.INFO

# Processing configuration
START_ROW = 2  # Skip header row
SAVE_FREQUENCY = 5  # Save file every N processed rows

# Resume configuration
RESUME_FILE = Path(".answer_generator_resume.json")  # Hidden file for resume state

# ============================================================================
# END OF CONFIGURATION SECTION
# ============================================================================

# Initialize module logger
logger = get_logger(__name__)

# Import required modules after logging setup
try:
    from embeddings import base_embedding
    from prompts import get_answer_prompt, get_category_prompt
    from storages import duckdb_qa_store
except ImportError as e:
    logger.error(f"Error: Could not import required modules: {e}")
    logger.error("Ensure all modules are in the correct paths:")
    logger.error("  - db/duckdb_qa_store.py")
    logger.error("  - embeddings/base_embedding.py")
    logger.error("  - prompts/get_category_prompt.py")
    logger.error("  - prompts/get_answer_prompt.py")
    sys.exit(1)


def format_json_for_excel(data: Any) -> str:
    """Format JSON data for Excel cell display.

    Formats JSON data (messages or raw responses) for readable display
    in Excel cells while preserving the complete data structure.

    Args:
        data: JSON-serializable data (dict, list, or already stringified JSON).

    Returns:
        Formatted JSON string suitable for Excel cell display.
    """
    try:
        # If already a string, try to parse and re-format
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                # Already formatted or not JSON, return as-is
                return str(data)  # Explicit string conversion

        # Format with indentation and ensure_ascii=False for readability
        formatted = json.dumps(data, ensure_ascii=False, indent=2)

        # Replace escaped newlines with actual newlines for Excel
        formatted = formatted.replace("\\n", "\n")

        return formatted

    except Exception as e:
        logger.warning(f"Could not format JSON for Excel: {e}")
        # Return string representation as fallback
        return str(data)


class AnswerGenerator:
    """Main class for generating answers using RAG approach with DuckDB."""

    def __init__(self):
        """Initialize the answer generator."""
        self.workbook: Workbook | None = None
        self.output_file: Path | None = None
        self.db_store: duckdb_qa_store.QADatabaseStore | None = None
        self.categories: dict[str, str] = {}
        self.conference_topic: str | None = None
        self.processed_count: int = 0
        self.resume_state: dict = {}
        self.timestamp: str = ""

    def extract_response_content(self, raw_response: Any) -> str:
        """Extract content field from LLM response.

        Args:
            raw_response: Raw response from LLM (dict or str).

        Returns:
            Content field value or empty string on error.
        """
        try:
            # Convert to dict if string
            if isinstance(raw_response, str):
                raw_response = json.loads(raw_response)

            # Navigate: choices[0].message.content
            content = raw_response.get("choices", [{}])[0].get("message", {}).get("content", "")
            # Ensure string return type
            return str(content) if content is not None else ""
        except (json.JSONDecodeError, KeyError, IndexError, TypeError, AttributeError):
            return ""

    def validate_input_files(self) -> bool:
        """Validate that required input files exist and are accessible.

        Returns:
            True if validation passes, False otherwise.
        """
        # Check Q.xlsx
        if not INPUT_FILE_Q.exists():
            logger.error(f"Input file not found: {INPUT_FILE_Q}")
            return False

        # Check if Q.xlsx is accessible
        try:
            with open(INPUT_FILE_Q, "rb") as f:
                pass
        except PermissionError:
            logger.error(f"Cannot access file: {INPUT_FILE_Q}")
            logger.error("The file might be open in Excel or another program.")
            return False
        except OSError as e:
            logger.error(f"Cannot read file {INPUT_FILE_Q}: {e}")
            return False

        # Check Q.xlsx structure
        try:
            wb = openpyxl.load_workbook(INPUT_FILE_Q, read_only=True)
            if SHEET_Q not in wb.sheetnames:
                logger.error(f"Required sheet '{SHEET_Q}' not found in {INPUT_FILE_Q}")
                wb.close()
                return False
            wb.close()
        except Exception as e:
            logger.error(f"Error reading {INPUT_FILE_Q}: {e}")
            return False

        # Check database file
        if not DATABASE_FILE.exists():
            logger.error(f"Database file not found: {DATABASE_FILE}")
            logger.error("Please run update_db.py first to create the database")
            return False

        return True

    def load_conference_topic(self, workbook: Workbook) -> str | None:
        """Load conference topic from T sheet if available.

        Args:
            workbook: The workbook to read from.

        Returns:
            Conference topic string or None if not available.
        """
        if SHEET_T not in workbook.sheetnames:
            logger.info(f"Sheet '{SHEET_T}' not found, proceeding without conference topic")
            return None

        sheet_t = workbook[SHEET_T]
        topic = sheet_t.cell(row=2, column=1).value  # A2 cell

        if topic and str(topic).strip():
            topic_str = str(topic).strip()
            logger.info(f"Conference topic loaded: {topic_str[:100]}...")
            return topic_str
        logger.info("No conference topic found in T.A2, proceeding without it")
        return None

    def load_categories_from_qa(self) -> dict[str, str]:
        """Load category dictionary from QA.xlsx if available.

        Returns:
            Dictionary mapping category names to descriptions, or empty dict.
        """
        if not INPUT_FILE_QA.exists():
            logger.info(f"File {INPUT_FILE_QA} not found, proceeding without categories")
            return {}

        try:
            wb = openpyxl.load_workbook(INPUT_FILE_QA, read_only=True)

            if SHEET_CATEGORY not in wb.sheetnames:
                logger.info(
                    f"Sheet '{SHEET_CATEGORY}' not found in {INPUT_FILE_QA}, proceeding without categories"
                )
                wb.close()
                return {}

            sheet_category = wb[SHEET_CATEGORY]
            categories = {}

            for row_idx in range(START_ROW, sheet_category.max_row + 1):
                category_name = sheet_category.cell(row=row_idx, column=COL_CATEGORY_NAME).value
                category_desc = sheet_category.cell(row=row_idx, column=COL_CATEGORY_DESC).value

                # Skip empty rows
                if category_name is None and category_desc is None:
                    continue

                # Validate completeness
                if category_name is None or category_desc is None:
                    logger.error(
                        f"Incomplete category data at row {row_idx}. "
                        f"Name: {category_name}, Description: {category_desc}"
                    )
                    wb.close()
                    raise ValueError("All categories must have both name and description")

                categories[str(category_name).strip()] = str(category_desc).strip()

            wb.close()

            if categories:
                logger.info(f"Loaded {len(categories)} categories from {INPUT_FILE_QA}")
            else:
                logger.info("No categories found in CATEGORY sheet")

            return categories

        except Exception as e:
            logger.error(f"Error loading categories: {e}")
            return {}

    def create_log_params_sheet(self, wb: Workbook) -> None:
        """Create or update LOG_ANSWER_PARAMS sheet with model parameters.

        Args:
            wb: Workbook to add the sheet to.
        """
        if not LOG_ANSWER:
            return

        # Remove existing sheet if present
        if SHEET_LOG_PARAMS in wb.sheetnames:
            wb.remove(wb[SHEET_LOG_PARAMS])

        # Create new sheet
        params_sheet = wb.create_sheet(SHEET_LOG_PARAMS)

        # Add headers
        params_sheet.cell(row=1, column=1, value="Parameter")
        params_sheet.cell(row=1, column=2, value="Value")
        params_sheet.cell(row=1, column=3, value="Description")

        # Configuration parameters with constant names
        row = 2
        config_params = [
            ("Timestamp", self.timestamp, "Processing start time"),
            ("TOP_K_SIMILAR", TOP_K_SIMILAR, "Number of similar questions to retrieve"),
            ("SIMILARITY_THRESHOLD", SIMILARITY_THRESHOLD, "Minimum cosine similarity"),
            (
                "CATEGORY_SEARCH_MODE",
                CATEGORY_SEARCH_MODE.value,
                "How categories are used in search",
            ),
            ("EMBEDDING_MODEL", EMBEDDING_MODEL, "Model for embeddings"),
            ("EMBEDDING_DIMENSION", EMBEDDING_DIMENSION, "Embedding vector size"),
            (
                "MAX_RETRY_ATTEMPTS_CATEGORY",
                MAX_RETRY_ATTEMPTS_CATEGORY,
                "Max retries for categorization",
            ),
            (
                "MAX_RETRY_ATTEMPTS_ANSWER",
                MAX_RETRY_ATTEMPTS_ANSWER,
                "Max retries for answer generation",
            ),
            ("RETRY_DELAY", RETRY_DELAY, "Delay between retries (seconds)"),
            ("SAVE_FREQUENCY", SAVE_FREQUENCY, "Save file every N rows"),
            ("LOG_ANSWER", LOG_ANSWER, "Enable detailed logging sheet"),
            ("LOG_TO_FILE", LOG_TO_FILE, "Enable logging to text file"),
            ("LOG_LEVEL", logging.getLevelName(LOG_LEVEL), "Logging detail level"),
            ("START_ROW", START_ROW, "First row to process (skip headers)"),
        ]

        for param_name, param_value, param_desc in config_params:
            params_sheet.cell(row=row, column=1, value=param_name)
            params_sheet.cell(row=row, column=2, value=str(param_value))
            params_sheet.cell(row=row, column=3, value=param_desc)
            row += 1

        # Add separator for category model parameters (always show if module exists)
        try:
            category_params = getattr(get_category_prompt, "params", {})
            if category_params:
                row += 1
                params_sheet.cell(row=row, column=1, value="--- Category Model Parameters ---")
                row += 1

                param_descriptions = {
                    "model": "LLM model name for categorization",
                    "temperature": "Randomness of responses (0-1)",
                    "top_p": "Nucleus sampling threshold",
                    "stream": "Whether to stream responses",
                    "max_tokens": "Maximum tokens in response",
                    "repetition_penalty": "Penalty for repeated tokens",
                }

                for param_name, param_value in category_params.items():
                    params_sheet.cell(row=row, column=1, value=f"category.{param_name}")
                    params_sheet.cell(row=row, column=2, value=str(param_value))
                    params_sheet.cell(
                        row=row, column=3, value=param_descriptions.get(param_name, "")
                    )
                    row += 1
        except AttributeError:
            logger.debug("Category prompt module params not available")

        # Add separator for answer model parameters
        try:
            answer_params = getattr(get_answer_prompt, "params", {})
            if answer_params:
                row += 1
                params_sheet.cell(row=row, column=1, value="--- Answer Model Parameters ---")
                row += 1

                param_descriptions = {
                    "model": "LLM model name for answer generation",
                    "temperature": "Randomness of responses (0-1)",
                    "top_p": "Nucleus sampling threshold",
                    "stream": "Whether to stream responses",
                    "max_tokens": "Maximum tokens in response",
                    "repetition_penalty": "Penalty for repeated tokens",
                }

                for param_name, param_value in answer_params.items():
                    params_sheet.cell(row=row, column=1, value=f"answer.{param_name}")
                    params_sheet.cell(row=row, column=2, value=str(param_value))
                    params_sheet.cell(
                        row=row, column=3, value=param_descriptions.get(param_name, "")
                    )
                    row += 1
        except AttributeError:
            logger.debug("Answer prompt module params not available")

        # Add separator for embedding model
        row += 1
        params_sheet.cell(row=row, column=1, value="--- Embedding Model ---")
        row += 1

        default_embedding_model = getattr(base_embedding, "DEFAULT_MODEL", "Unknown")
        params_sheet.cell(row=row, column=1, value="base_embedding.DEFAULT_MODEL")
        params_sheet.cell(row=row, column=2, value=default_embedding_model)
        params_sheet.cell(row=row, column=3, value="Default model for creating embeddings")

        # Auto-adjust column widths
        for column in params_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            params_sheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Created '{SHEET_LOG_PARAMS}' sheet with model parameters")

    def create_output_file(self) -> tuple[Workbook, Path]:
        """Create output file with timestamp.

        Returns:
            Tuple of (workbook, output_path).
        """
        # Ensure output directory exists
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Generate timestamp
        self.timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_file = OUTPUT_DIR / f"Q_{self.timestamp}.xlsx"

        # Setup file logging for this session
        if LOG_TO_FILE:
            log_file = OUTPUT_DIR / f"Q_{self.timestamp}.log"
            setup_logging(log_file=log_file, level=LOG_LEVEL)

        logger.info(f"Copying {INPUT_FILE_Q} to {output_file}")

        # Copy file to preserve structure
        shutil.copy2(INPUT_FILE_Q, output_file)
        logger.info("File copied successfully")

        # Open the copied file for modifications
        wb = openpyxl.load_workbook(output_file)

        # Add headers to Q sheet if not present
        sheet_q = wb[SHEET_Q]

        # Check if headers exist (assuming row 1 is header row)
        if sheet_q.cell(row=1, column=COL_Q_ANSWER).value is None:
            # Add headers for Q sheet with new columns
            q_headers = [
                "question",
                "answer",
                "answer_confidence",
                "answer_sources_used",
            ]
            # Add headers for similar Q&A pairs with similarity scores
            for i in range(1, TOP_K_SIMILAR + 1):
                q_headers.extend([f"s_{i}", f"q_{i}", f"a_{i}"])

            for col_idx, header in enumerate(q_headers, 1):
                sheet_q.cell(row=1, column=col_idx, value=header)

            logger.info(f"Added headers to '{SHEET_Q}' sheet")

        # Handle LOG_ANSWER sheet based on configuration
        log_sheet_name = SHEET_LOG_PREFIX

        if LOG_ANSWER:
            # Remove existing LOG_ANSWER sheet if it exists
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing existing '{log_sheet_name}' sheet for fresh start")
                wb.remove(wb[log_sheet_name])

            # Create new LOG_ANSWER sheet
            log_sheet = wb.create_sheet(log_sheet_name)

            # Add headers to LOG_ANSWER sheet with raw response and content columns
            headers = [
                "question",
                "category",
                "category_confidence",
                "category_reasoning",
                "category_messages",
                "category_response",
                "category_response_content",  # New column
                "answer",
                "answer_confidence",
                "answer_sources_used",
                "answer_messages",
                "answer_response",
                "answer_response_content",  # New column
            ]
            for col_idx, header in enumerate(headers, 1):
                log_sheet.cell(row=1, column=col_idx, value=header)

            logger.info(
                f"Created '{log_sheet_name}' sheet with headers including response content columns"
            )

            # Create LOG_ANSWER_PARAMS sheet
            self.create_log_params_sheet(wb)
        else:
            # If LOG_ANSWER is False and sheet exists, remove it
            if log_sheet_name in wb.sheetnames:
                logger.info(f"Removing '{log_sheet_name}' sheet (LOG_ANSWER=False)")
                wb.remove(wb[log_sheet_name])
            if SHEET_LOG_PARAMS in wb.sheetnames:
                logger.info(f"Removing '{SHEET_LOG_PARAMS}' sheet (LOG_ANSWER=False)")
                wb.remove(wb[SHEET_LOG_PARAMS])

        # Save modifications
        wb.save(output_file)
        logger.info(f"Created output file: {output_file}")

        return wb, output_file

    def save_resume_state(self, output_file: Path, last_row: int) -> None:
        """Save resume state to file.

        Args:
            output_file: Path to the output file being processed.
            last_row: Last successfully processed row number.
        """
        state = {
            "output_file": str(output_file),
            "last_row": last_row,
            "timestamp": datetime.now().isoformat(),
        }

        try:
            with open(RESUME_FILE, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2, ensure_ascii=False)
            logger.debug(f"Saved resume state at row {last_row}")
        except Exception as e:
            logger.warning(f"Could not save resume state: {e}")

    def load_resume_state(self) -> dict[Any, Any] | None:
        """Load resume state if exists.

        Returns:
            Resume state dictionary or None if not found.
        """
        if not RESUME_FILE.exists():
            return None

        try:
            with open(RESUME_FILE, encoding="utf-8") as f:
                state: dict[Any, Any] = json.load(f)  # Explicit type annotation

            # Validate output file still exists
            output_path = Path(state["output_file"])
            if output_path.exists():
                logger.info(f"Found resume state from {state['timestamp']}")
                logger.info(f"Will continue from row {state['last_row'] + 1}")
                return state
            logger.warning("Resume output file not found, starting fresh")
            return None

        except Exception as e:
            logger.warning(f"Could not load resume state: {e}")
            return None

    def clear_resume_state(self) -> None:
        """Clear resume state file."""
        if RESUME_FILE.exists():
            try:
                RESUME_FILE.unlink()
                logger.debug("Cleared resume state")
            except Exception as e:
                logger.warning(f"Could not clear resume state: {e}")

    def categorize_question_with_retry(self, question: str) -> dict[str, Any] | None:
        """Categorize a question with retry logic on errors.

        Args:
            question: The question to categorize.

        Returns:
            Categorization result dictionary or None if no categories.
        """
        if not self.categories:
            return None

        last_error = None

        for attempt in range(1, MAX_RETRY_ATTEMPTS_CATEGORY + 1):
            try:
                # Check if get_category_prompt.run() returns 2 or 3 values
                result = get_category_prompt.run(question)

                # Handle both old (2-tuple) and new (3-tuple) return formats
                if isinstance(result, tuple):
                    if len(result) == 2:
                        result_json, messages_list = result
                        raw_response = None
                    elif len(result) == 3:
                        result_json, messages_list, raw_response = result
                    else:
                        raise ValueError(
                            f"Unexpected return format from get_category_prompt.run: {len(result)} values"
                        )
                else:
                    raise ValueError(
                        f"Unexpected return type from get_category_prompt.run: {type(result)}"
                    )

                result_dict: dict[str, Any] = json.loads(result_json)  # Explicit type

                # Format messages for logging
                result_dict["messages"] = format_json_for_excel(messages_list)

                # Add raw response if available
                if raw_response is not None:
                    result_dict["response"] = format_json_for_excel(raw_response)
                    result_dict["response_content"] = self.extract_response_content(raw_response)
                else:
                    result_dict["response"] = ""
                    result_dict["response_content"] = ""

                if attempt > 1:
                    logger.info(f"Categorization succeeded on attempt {attempt}")

                return result_dict

            except Exception as e:
                last_error = e
                logger.warning(
                    f"Categorization attempt {attempt}/{MAX_RETRY_ATTEMPTS_CATEGORY} failed: {e}"
                )

                if attempt < MAX_RETRY_ATTEMPTS_CATEGORY:
                    if RETRY_DELAY > 0:
                        time.sleep(RETRY_DELAY)
                    continue

        # All attempts failed
        logger.error(f"All {MAX_RETRY_ATTEMPTS_CATEGORY} categorization attempts failed")
        return {
            "category": "Error",
            "confidence": 0.0,
            "reasoning": f"All retry attempts failed. Last error: {last_error!s}",
            "messages": "[]",
            "response": "",
            "response_content": "",
        }

    def search_similar_questions(
        self, question: str, category: str | None = None
    ) -> list[dict[str, Any]]:
        """Search for similar questions in the database.

        Uses CATEGORY_SEARCH_MODE to determine how to handle category filtering:
        - WITH_CATEGORY: Use category if provided
        - WITHOUT_CATEGORY: Always ignore category
        - CATEGORY_FALLBACK: Try with category first, then without if no results

        Args:
            question: The question to search for.
            category: Optional category filter.

        Returns:
            List of similar Q&A pairs with similarity scores.
        """
        if not self.db_store:
            logger.error("Database store not initialized")
            return []

        try:
            # Get question embedding
            embeddings, _ = base_embedding.create_embeddings(
                question, model=EMBEDDING_MODEL, task_type="query"
            )

            if not embeddings or not embeddings[0]:
                logger.error("Failed to create embedding for question")
                return []

            # Determine search strategy based on mode
            if CATEGORY_SEARCH_MODE == CategorySearchMode.WITHOUT_CATEGORY:
                # Always search without category
                logger.debug("Searching without category filter (mode: WITHOUT_CATEGORY)")
                results = self.db_store.search_similar_questions(
                    question_embedding=embeddings[0],
                    category=None,
                    top_k=TOP_K_SIMILAR,
                    threshold=SIMILARITY_THRESHOLD,
                )
            elif CATEGORY_SEARCH_MODE == CategorySearchMode.WITH_CATEGORY:
                # Use category if available
                if category:
                    logger.debug(f"Searching with category filter: {category}")
                else:
                    logger.debug("No category available, searching without filter")
                results = self.db_store.search_similar_questions(
                    question_embedding=embeddings[0],
                    category=category,
                    top_k=TOP_K_SIMILAR,
                    threshold=SIMILARITY_THRESHOLD,
                )
            # Try with category first, then without if no results
            elif category:
                logger.debug(f"Trying search with category filter: {category}")
                results = self.db_store.search_similar_questions(
                    question_embedding=embeddings[0],
                    category=category,
                    top_k=TOP_K_SIMILAR,
                    threshold=SIMILARITY_THRESHOLD,
                )

                if not results:
                    logger.info("No results with category, falling back to search without filter")
                    results = self.db_store.search_similar_questions(
                        question_embedding=embeddings[0],
                        category=None,
                        top_k=TOP_K_SIMILAR,
                        threshold=SIMILARITY_THRESHOLD,
                    )
            else:
                # No category to use, search without filter
                logger.debug("No category available, searching without filter")
                results = self.db_store.search_similar_questions(
                    question_embedding=embeddings[0],
                    category=None,
                    top_k=TOP_K_SIMILAR,
                    threshold=SIMILARITY_THRESHOLD,
                )

            return results

        except Exception as e:
            logger.error(f"Search failed: {e}")
            return []

    def generate_answer_with_retry(
        self, question: str, qa_pairs: list[dict[str, str]]
    ) -> dict[str, Any] | None:
        """Generate answer with retry logic on errors.

        Args:
            question: The question to answer.
            qa_pairs: Context Q&A pairs.

        Returns:
            Answer result dictionary or None on failure.
        """
        last_error = None

        for attempt in range(1, MAX_RETRY_ATTEMPTS_ANSWER + 1):
            try:
                # get_answer_prompt.run now returns 3 values
                result_json, messages_list, raw_response = get_answer_prompt.run(
                    user_question=question, qa_pairs=qa_pairs
                )
                result: dict[str, Any] = json.loads(result_json)  # Explicit type

                # Format messages for logging
                result["messages"] = format_json_for_excel(messages_list)

                # Format raw response for logging
                result["response"] = format_json_for_excel(raw_response)

                # Extract content from response
                result["response_content"] = self.extract_response_content(raw_response)

                if attempt > 1:
                    logger.info(f"Answer generation succeeded on attempt {attempt}")

                return result

            except Exception as e:
                last_error = e
                logger.warning(
                    f"Answer generation attempt {attempt}/{MAX_RETRY_ATTEMPTS_ANSWER} failed: {e}"
                )

                if attempt < MAX_RETRY_ATTEMPTS_ANSWER:
                    if RETRY_DELAY > 0:
                        time.sleep(RETRY_DELAY)
                    continue

        # All attempts failed
        logger.error(f"All {MAX_RETRY_ATTEMPTS_ANSWER} answer generation attempts failed")
        return {
            "answer": f"Failed to generate answer after {MAX_RETRY_ATTEMPTS_ANSWER} attempts. Last error: {last_error!s}",
            "confidence": 0.0,
            "sources_used": [],
            "messages": "[]",
            "response": "",
            "response_content": "",
        }

    def process_row(self, sheet_q: Worksheet, sheet_log: Worksheet | None, row_idx: int) -> bool:
        """Process a single row from Q sheet.

        Args:
            sheet_q: The Q worksheet.
            sheet_log: The LOG_ANSWER worksheet (if enabled).
            row_idx: Row index to process.

        Returns:
            True if processing was successful, False otherwise.
        """
        try:
            # Get question from Q sheet
            question = sheet_q.cell(row=row_idx, column=COL_Q_QUESTION).value

            if question is None:
                logger.warning(f"Empty question at row {row_idx}, skipping")
                return True

            question_str = str(question).strip()
            question_preview = question_str[:80] + "..." if len(question_str) > 80 else question_str
            logger.info(f"Processing row {row_idx}: {question_preview}")

            # Step 1: Categorize question (optional) with retry
            category_result = None
            selected_category = None

            if self.categories:
                category_result = self.categorize_question_with_retry(question_str)
                if category_result:
                    selected_category = category_result.get("category")
                    if selected_category != "Error":
                        logger.info(
                            f"Categorized as: {selected_category} "
                            f"(confidence: {category_result.get('confidence', 0):.2f})"
                        )

            # Step 2: Search similar questions
            similar_questions = self.search_similar_questions(question_str, selected_category)

            if not similar_questions:
                logger.warning(f"No similar questions found for row {row_idx}")
                # Write empty result
                sheet_q.cell(row=row_idx, column=COL_Q_ANSWER, value="No similar questions found")
                sheet_q.cell(row=row_idx, column=COL_Q_CONFIDENCE, value=0.0)
                sheet_q.cell(row=row_idx, column=COL_Q_SOURCES, value="[]")
                return True

            logger.info(f"Found {len(similar_questions)} similar questions")

            # Step 3: Prepare context for answer generation
            qa_pairs = [
                {"question": result["question"], "answer": result["answer"]}
                for result in similar_questions
            ]

            # Step 4: Generate answer with retry
            answer_result = self.generate_answer_with_retry(question_str, qa_pairs)

            if not answer_result:
                logger.error(f"Failed to generate answer for row {row_idx}")
                sheet_q.cell(row=row_idx, column=COL_Q_ANSWER, value="Failed to generate answer")
                sheet_q.cell(row=row_idx, column=COL_Q_CONFIDENCE, value=0.0)
                sheet_q.cell(row=row_idx, column=COL_Q_SOURCES, value="[]")
                return False

            # Step 5: Write results to Q sheet with new columns
            # Write answer and its metadata
            sheet_q.cell(row=row_idx, column=COL_Q_ANSWER, value=answer_result.get("answer", ""))
            sheet_q.cell(
                row=row_idx,
                column=COL_Q_CONFIDENCE,
                value=answer_result.get("confidence", 0.0),
            )
            # Format sources_used as string for Excel
            sources_used = answer_result.get("sources_used", [])
            if isinstance(sources_used, list):
                sources_str = json.dumps(sources_used, ensure_ascii=False)
            else:
                sources_str = str(sources_used)
            sheet_q.cell(row=row_idx, column=COL_Q_SOURCES, value=sources_str)

            # Write similar Q&A pairs with similarity scores
            for i, result in enumerate(similar_questions[:TOP_K_SIMILAR]):
                # Calculate column positions: s_i, q_i, a_i
                s_col = COL_Q_S1 + (i * 3)  # s_1, s_2, s_3, etc.
                q_col = COL_Q_Q1 + (i * 3)  # q_1, q_2, q_3, etc.
                a_col = COL_Q_A1 + (i * 3)  # a_1, a_2, a_3, etc.

                # Write similarity score
                similarity_score = result.get("similarity", 0.0)
                sheet_q.cell(row=row_idx, column=s_col, value=round(similarity_score, 4))
                # Write question and answer
                sheet_q.cell(row=row_idx, column=q_col, value=result["question"])
                sheet_q.cell(row=row_idx, column=a_col, value=result["answer"])

            # Step 6: Write to LOG_ANSWER sheet if enabled (with raw responses and content)
            if LOG_ANSWER and sheet_log:
                col = 1
                sheet_log.cell(row=row_idx, column=col, value=question_str)
                col += 1

                if category_result:
                    sheet_log.cell(
                        row=row_idx,
                        column=col,
                        value=category_result.get("category", ""),
                    )
                    col += 1
                    sheet_log.cell(
                        row=row_idx,
                        column=col,
                        value=category_result.get("confidence", ""),
                    )
                    col += 1
                    sheet_log.cell(
                        row=row_idx,
                        column=col,
                        value=category_result.get("reasoning", ""),
                    )
                    col += 1
                    sheet_log.cell(
                        row=row_idx,
                        column=col,
                        value=category_result.get("messages", ""),
                    )
                    col += 1
                    sheet_log.cell(
                        row=row_idx,
                        column=col,
                        value=category_result.get("response", ""),
                    )
                    col += 1
                    sheet_log.cell(
                        row=row_idx,
                        column=col,
                        value=category_result.get("response_content", ""),
                    )
                    col += 1
                else:
                    # Skip category columns if no categorization (6 columns total)
                    col += 6

                sheet_log.cell(row=row_idx, column=col, value=answer_result.get("answer", ""))
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=answer_result.get("confidence", ""))
                col += 1
                sheet_log.cell(
                    row=row_idx,
                    column=col,
                    value=str(answer_result.get("sources_used", [])),
                )
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=answer_result.get("messages", ""))
                col += 1
                sheet_log.cell(row=row_idx, column=col, value=answer_result.get("response", ""))
                col += 1
                sheet_log.cell(
                    row=row_idx,
                    column=col,
                    value=answer_result.get("response_content", ""),
                )

            logger.info(f"Row {row_idx} processed successfully")
            self.processed_count += 1
            return True

        except Exception as e:
            logger.error(f"Error processing row {row_idx}: {e}")
            return False

    def run(self) -> bool:
        """Main execution method.

        Returns:
            True if execution completed successfully, False otherwise.
        """
        try:
            logger.info("=" * 70)
            logger.info("Starting Answer Generator Script")
            logger.info(f"Log Answer: {'ENABLED' if LOG_ANSWER else 'DISABLED'}")
            logger.info(f"Category Search Mode: {CATEGORY_SEARCH_MODE.value}")
            logger.info(f"Max Retry Attempts (Category): {MAX_RETRY_ATTEMPTS_CATEGORY}")
            logger.info(f"Max Retry Attempts (Answer): {MAX_RETRY_ATTEMPTS_ANSWER}")
            logger.info("=" * 70)

            # Check for resume state
            resume_state = self.load_resume_state()

            if resume_state:
                # Resume from previous run
                logger.info("Resuming from previous run...")
                self.output_file = Path(resume_state["output_file"])

                # Extract timestamp from filename for log file
                filename = self.output_file.stem  # Q_YYYY-MM-DD_HHMMSS
                if filename.startswith("Q_"):
                    self.timestamp = filename[2:]  # Remove "Q_" prefix

                # Setup file logging for resumed session
                if LOG_TO_FILE:
                    log_file = OUTPUT_DIR / f"Q_{self.timestamp}.log"
                    setup_logging(log_file=log_file, level=LOG_LEVEL)
                    logger.info("=" * 70)
                    logger.info("RESUMED SESSION")
                    logger.info("=" * 70)

                self.workbook = openpyxl.load_workbook(self.output_file)
                start_from_row = resume_state["last_row"] + 1
            else:
                # Fresh start
                logger.info("Starting fresh processing...")

                # Step 1: Validate input files
                logger.info("Step 1: Validating input files...")
                if not self.validate_input_files():
                    logger.error("=" * 70)
                    logger.error("FAILED: Cannot proceed with processing")
                    logger.error("=" * 70)
                    return False

                # Step 2: Create output file (this also sets up file logging)
                logger.info("Step 2: Creating output file...")
                self.workbook, self.output_file = self.create_output_file()

                start_from_row = START_ROW

            # Step 3: Load conference topic
            logger.info("Step 3: Loading conference topic...")
            self.conference_topic = self.load_conference_topic(self.workbook)

            # Step 4: Load categories
            logger.info("Step 4: Loading categories...")
            self.categories = self.load_categories_from_qa()

            if self.categories:
                logger.info(f"Categories loaded: {', '.join(self.categories.keys())}")
                # Initialize category prompt
                get_category_prompt.update_system_prompt(categories=self.categories)
                logger.info("Category system initialized")

            # Step 5: Initialize database connection
            logger.info("Step 5: Connecting to database...")
            self.db_store = duckdb_qa_store.QADatabaseStore(
                str(DATABASE_FILE), embedding_size=EMBEDDING_DIMENSION
            )
            logger.info("Database connected successfully")

            # Step 6: Initialize answer generation system
            logger.info("Step 6: Initializing answer generation system...")
            get_answer_prompt.update_system_prompt(topic=self.conference_topic)
            logger.info("Answer generation system initialized")

            # Step 7: Process questions
            logger.info("Step 7: Processing questions...")

            sheet_q = self.workbook[SHEET_Q]

            # Get LOG_ANSWER sheet if enabled
            sheet_log = None
            if LOG_ANSWER:
                if SHEET_LOG_PREFIX in self.workbook.sheetnames:
                    sheet_log = self.workbook[SHEET_LOG_PREFIX]
                else:
                    logger.warning(
                        f"LOG_ANSWER is enabled but '{SHEET_LOG_PREFIX}' sheet not found"
                    )

            # Determine rows to process
            rows_to_process = []
            for row_idx in range(start_from_row, sheet_q.max_row + 1):
                question = sheet_q.cell(row=row_idx, column=COL_Q_QUESTION).value
                if question and str(question).strip():
                    rows_to_process.append(row_idx)

            if not rows_to_process:
                logger.info("No questions to process")
                self.clear_resume_state()
                return True

            logger.info(f"Found {len(rows_to_process)} questions to process")
            logger.info("-" * 70)

            # Process each row
            successfully_processed: list[int] = []

            for idx, row_idx in enumerate(rows_to_process, 1):
                # Process row
                success = self.process_row(sheet_q, sheet_log, row_idx)

                if not success:
                    logger.error(f"Failed to process row {row_idx}")
                    # Save progress on error
                    if successfully_processed:
                        try:
                            self.workbook.save(self.output_file)
                            last_successful = successfully_processed[-1]
                            self.save_resume_state(self.output_file, last_successful)
                            logger.info(
                                f"Emergency save after error: saved up to row {last_successful}"
                            )
                        except Exception as save_error:
                            logger.error(f"Could not perform emergency save: {save_error}")
                    continue

                # Track successful processing
                successfully_processed.append(row_idx)

                # Periodic save
                if idx % SAVE_FREQUENCY == 0:
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(f"Progress saved: {idx}/{len(rows_to_process)} rows processed")
                    except Exception as save_error:
                        logger.error(f"Could not save progress: {save_error}")

                # Final save
                if idx == len(rows_to_process):
                    try:
                        self.workbook.save(self.output_file)
                        self.save_resume_state(self.output_file, row_idx)
                        logger.info(
                            f"Final batch saved: {idx}/{len(rows_to_process)} rows completed"
                        )
                    except Exception as save_error:
                        logger.error(f"Could not save final batch: {save_error}")

            # Final save
            self.workbook.save(self.output_file)
            logger.info(f"Final save completed: {self.output_file}")

            # Clear resume state on successful completion
            self.clear_resume_state()

            # Close database connection
            if self.db_store:
                self.db_store.close()

            # Summary
            logger.info("-" * 70)
            logger.info("Processing completed successfully!")
            logger.info(f"Total rows processed: {self.processed_count}")
            logger.info(f"Output file: {self.output_file}")
            if LOG_ANSWER and sheet_log:
                logger.info(f"Detailed logs saved in '{SHEET_LOG_PREFIX}' sheet")
                logger.info(f"Model parameters saved in '{SHEET_LOG_PARAMS}' sheet")
            if LOG_TO_FILE:
                log_path = OUTPUT_DIR / f"Q_{self.timestamp}.log"
                logger.info(f"Text log saved to: {log_path}")
            logger.info("=" * 70)

            return True

        except KeyboardInterrupt:
            logger.warning("Processing interrupted by user (Ctrl+C)")
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Progress saved to: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as e:
                    logger.error(f"Could not save progress: {e}")
            return False

        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)

            # Try to save current state
            if self.workbook and self.output_file:
                try:
                    self.workbook.save(self.output_file)
                    logger.info(f"Emergency save completed: {self.output_file}")
                    logger.info("You can resume processing by running the script again")
                except Exception as save_error:
                    logger.error(f"Could not save file: {save_error}")

            return False

        finally:
            # Close workbook
            if self.workbook:
                try:
                    self.workbook.close()
                except Exception:
                    pass

            # Close database
            if self.db_store:
                try:
                    self.db_store.close()
                except Exception:
                    pass


def main():
    """Main entry point for the script."""
    try:
        # Initialize logging for the main application
        setup_logging(level=LOG_LEVEL)

        generator = AnswerGenerator()
        success = generator.run()

        if success:
            logger.info("Script completed successfully")
        else:
            logger.error("Script completed with errors")

        # Exit with appropriate code
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("\nScript interrupted by user")
        sys.exit(130)  # Standard exit code for Ctrl+C
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)
    finally:
        # Clean up logging resources
        close_logging()


if __name__ == "__main__":
    main()
